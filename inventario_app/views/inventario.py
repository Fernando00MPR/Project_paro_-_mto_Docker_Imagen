from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.db.models import Count
import openpyxl
from paros_app.views.utils import _excel_response, _estilo_cabecera
from django.views.decorators.http import require_POST

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io

from mto_app.models import Area
from ..models import Refaccion, CategoriaRefaccion, ImagenRefaccion


@login_required
def lista_refacciones(request):
    acceso = getattr(request.user, 'acceso_mto', None)
    puede_ver = (
        request.user.is_superuser or
        (hasattr(request.user, 'perfil') and request.user.perfil.es_admin) or
        (acceso and acceso.ver_inventario)
    )
    if not puede_ver:
        messages.error(request, "No tienes permiso para ver esta sección.")
        return redirect('mto:dashboard')

    area_id       = request.GET.get('area', '')
    categoria_id  = request.GET.get('categoria', '')
    busqueda      = request.GET.get('q', '').strip()
    estatus_stock = request.GET.get('estatus_stock', '')
    criticidad_id = request.GET.get('criticidad', '')

    refacciones_qs = Refaccion.objects.select_related('area', 'categoria').annotate(
        num_imagenes=Count('imagenes')
    ).filter(activo=True)

    if area_id:
        refacciones_qs = refacciones_qs.filter(area_id=area_id)
    if categoria_id:
        refacciones_qs = refacciones_qs.filter(categoria_id=categoria_id)
    if criticidad_id:
        refacciones_qs = refacciones_qs.filter(criticidad=criticidad_id)
    if busqueda:
        refacciones_qs = (refacciones_qs.filter(nombre__icontains=busqueda)  |
                          refacciones_qs.filter(no_item__icontains=busqueda) |
                          refacciones_qs.filter(descripcion__icontains=busqueda))
        
    refacciones_lista = list(refacciones_qs)
    total_bajo_minimo = sum(1 for r in refacciones_lista if r.bajo_minimo)

    if estatus_stock == 'bajo_minimo':
        refacciones = [r for r in refacciones_lista if r.bajo_minimo]
    elif estatus_stock == 'sobre_maximo':
        refacciones = [r for r in refacciones_lista if r.sobre_maximo]
    elif estatus_stock == 'en_rango':
        refacciones = [r for r in refacciones_lista if not r.bajo_minimo and not r.sobre_maximo and r.stock_actual > 0]
    elif estatus_stock == 'sin_stock':
        refacciones = [r for r in refacciones_lista if r.stock_actual == 0]
    else:
        refacciones = refacciones_lista

    per_page = request.GET.get('per_page', '10')
    paginator = Paginator(refacciones, int(per_page) if per_page.isdigit() else 10)
    page_num = request.GET.get('page', 1)
    refacciones_page = paginator.get_page(page_num)

    ctx = {
        'refacciones':       refacciones_page,
        'areas':             Area.objects.filter(activa=True),
        'categorias':        CategoriaRefaccion.objects.all(),
        'filtro_area':       area_id,
        'filtro_criticidad': criticidad_id,
        'filtro_categoria':  categoria_id,
        'busqueda':          busqueda,
        'estatus_stock':     estatus_stock,
        'per_page':          per_page,
        'total_bajo_minimo': total_bajo_minimo,
        'unidades':          Refaccion.UNIDAD_CHOICES,
        'criticidades':      Refaccion.CRITICIDAD_CHOICES,
        'puede_editar_inventario':   (request.user.is_superuser or (hasattr(request.user, 'perfil') and request.user.perfil.es_admin) or (acceso and acceso.editar_inventario)),
        'puede_eliminar_inventario': (request.user.is_superuser or (hasattr(request.user, 'perfil') and request.user.perfil.es_admin) or (acceso and acceso.eliminar_inventario)),
    }
    return render(request, 'inventario_app/lista_refacciones.html', ctx)


@login_required
def form_refaccion(request, pk=None):
    acceso = getattr(request.user, 'acceso_mto', None)
    puede_editar = (
        request.user.is_superuser or
        (hasattr(request.user, 'perfil') and request.user.perfil.es_admin) or
        (acceso and acceso.editar_inventario)
    )
    if not puede_editar:
        messages.error(request, "No tienes permiso para editar el inventario.")
        return redirect('inventario:lista_refacciones')

    refaccion = get_object_or_404(Refaccion, pk=pk) if pk else None

    if request.method == 'POST':
        try:
            datos = {
                'no_item':        request.POST.get('no_item', '').strip(),
                'nombre':         request.POST.get('nombre', '').strip(),
                'descripcion':    request.POST.get('descripcion', '').strip(),
                'area':           get_object_or_404(Area, pk=request.POST.get('area')),
                'criticidad':     request.POST.get('criticidad', 'no_critico'),
                'categoria_id':   request.POST.get('categoria') or None,
                'unidad':         request.POST.get('unidad', 'pza'),
                'stock_actual':   int(request.POST.get('stock_actual', 0) or 0),
                'stock_minimo':   int(request.POST.get('stock_minimo', 0) or 0),
                'stock_maximo':   int(request.POST.get('stock_maximo', 0) or 0),
                'ubicacion':      request.POST.get('ubicacion', '').strip(),
                'proveedor':      request.POST.get('proveedor', '').strip(),
                'costo_unitario': request.POST.get('costo_unitario') or None,
                'activo':         'activo' in request.POST,
            }

            if not datos['no_item'] or not datos['nombre']:
                raise ValueError("No. Item y Nombre son obligatorios.")

            area_pk = datos['area'].pk

            if refaccion:
                for k, v in datos.items():
                    setattr(refaccion, k, v)
                refaccion.save()
                messages.success(request, f"Refacción '{refaccion.no_item}' actualizada.")
            else:
                refaccion = Refaccion.objects.create(**datos)
                messages.success(request, "Refacción creada.")

            imagenes_nuevas = request.FILES.getlist('imagenes')
            existentes      = refaccion.imagenes.count()
            disponibles     = 2 - existentes

            if imagenes_nuevas and disponibles <= 0:
                messages.warning(request, "Ya se alcanzó el máximo de 2 imágenes; no se agregaron las nuevas.")
            else:
                for imagen in imagenes_nuevas[:max(disponibles, 0)]:
                    ImagenRefaccion.objects.create(refaccion=refaccion, imagen=imagen)
                if len(imagenes_nuevas) > disponibles:
                    messages.warning(request, f"Solo se guardaron {max(disponibles,0)} imagen(es); límite de 2 alcanzado.")

            return redirect(f"{reverse('inventario:lista_refacciones')}?area={area_pk}")

        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f"Error al guardar: {e}")

        area_id = request.POST.get('area', '')
        return redirect(f"{reverse('inventario:lista_refacciones')}?area={area_id}")

    return redirect('inventario:lista_refacciones')


@login_required
def eliminar_refaccion(request, pk):
    acceso = getattr(request.user, 'acceso_mto', None)
    puede_eliminar = (
        request.user.is_superuser or
        (hasattr(request.user, 'perfil') and request.user.perfil.es_admin) or
        (acceso and acceso.eliminar_inventario)
    )
    if not puede_eliminar:
        messages.error(request, "No tienes permiso para eliminar refacciones.")
        return redirect('inventario:lista_refacciones')

    refaccion = get_object_or_404(Refaccion, pk=pk)
    if request.method == 'POST':
        cantidad_seguimientos = refaccion.seguimientos.count()
        if cantidad_seguimientos > 0:
            messages.error(
                request,
                f"No se puede eliminar '{refaccion.no_item}' porque tiene "
                f"{cantidad_seguimientos} seguimiento(s) de compra asociado(s). "
                f"Elimina primero esos seguimientos."
            )
        else:
            refaccion.delete()
            messages.success(request, f"Refacción '{refaccion.no_item}' eliminada.")
    return redirect('inventario:lista_refacciones')


@login_required
@require_POST
def subir_imagenes_refaccion(request, refaccion_pk):
    acceso = getattr(request.user, 'acceso_mto', None)
    puede_editar = (
        request.user.is_superuser or
        (hasattr(request.user, 'perfil') and request.user.perfil.es_admin) or
        (acceso and acceso.editar_inventario)
    )
    if not puede_editar:
        return JsonResponse({'ok': False, 'error': 'Sin permiso'}, status=403)

    refaccion = get_object_or_404(Refaccion, pk=refaccion_pk)
    imagenes  = request.FILES.getlist('imagenes')
    existentes  = refaccion.imagenes.count()
    disponibles = 2 - existentes

    creadas = []
    for imagen in imagenes[:max(disponibles, 0)]:
        img = ImagenRefaccion.objects.create(refaccion=refaccion, imagen=imagen)
        creadas.append({'id': img.id, 'url': img.imagen.url})

    return JsonResponse({'ok': True, 'imagenes': creadas, 'total': refaccion.imagenes.count()})


@login_required
def imagenes_refaccion(request, refaccion_pk):
    refaccion = get_object_or_404(Refaccion, pk=refaccion_pk)
    imagenes = [{'id': img.id, 'url': img.imagen.url} for img in refaccion.imagenes.all()]
    return JsonResponse({'imagenes': imagenes})


@login_required
@require_POST
def eliminar_imagen_refaccion(request, imagen_id):
    acceso = getattr(request.user, 'acceso_mto', None)
    puede_editar = (
        request.user.is_superuser or
        (hasattr(request.user, 'perfil') and request.user.perfil.es_admin) or
        (acceso and acceso.editar_inventario)
    )
    if not puede_editar:
        return JsonResponse({'ok': False, 'error': 'Sin permiso'}, status=403)

    imagen = get_object_or_404(ImagenRefaccion, id=imagen_id)
    imagen.imagen.delete()
    imagen.delete()
    return JsonResponse({'ok': True})


@login_required
def buscar_refacciones(request):
    q = request.GET.get('q', '').strip()
    area_id = request.GET.get('area', '')

    qs = Refaccion.objects.filter(activo=True)
    if area_id:
        qs = qs.filter(area_id=area_id)
    if q:
        qs = (qs.filter(no_item__icontains=q) | qs.filter(nombre__icontains=q))

    resultados = [
        {'id': r.id, 'no_item': r.no_item, 'nombre': r.nombre}
        for r in qs[:15]
    ]
    return JsonResponse(resultados, safe=False)


@login_required
def importar_stock(request):
    acceso = getattr(request.user, 'acceso_mto', None)
    puede_editar = (
        request.user.is_superuser or
        (hasattr(request.user, 'perfil') and request.user.perfil.es_admin) or
        (acceso and acceso.editar_inventario)
    )
    if not puede_editar:
        messages.error(request, "No tienes permiso para importar inventario.")
        return redirect('inventario:lista_refacciones')

    area_id = request.GET.get('area', '') or request.POST.get('area', '')

    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        area_id = request.POST.get('area', '').strip()

        if not archivo or not area_id:
            messages.error(request, "Debes seleccionar un área y subir un archivo.")
            return redirect(f"{reverse('inventario:importar_stock')}?area={area_id}")

        try:
            area = get_object_or_404(Area, pk=area_id)
            wb = load_workbook(archivo, data_only=True)
            ws = wb.active

            actualizados = 0
            ignorados = 0

            refacciones_dict = {r.no_item: r for r in Refaccion.objects.filter(area=area)}
            a_actualizar = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                no_item = row[0]
                stock_actual = row[2]

                if no_item is None or stock_actual is None:
                    continue

                no_item = str(no_item).strip()
                if not no_item:
                    continue

                try:
                    # Acepta int, float o string numérico; siempre lo convierte a entero
                    stock_actual = int(float(str(stock_actual).strip()))
                except (ValueError, TypeError):
                    ignorados += 1
                    continue

                refaccion = refacciones_dict.get(no_item)
                if refaccion:
                    refaccion.stock_actual = stock_actual
                    a_actualizar.append(refaccion)
                    actualizados += 1
                else:
                    ignorados += 1
            if a_actualizar:
                Refaccion.objects.bulk_update(a_actualizar, ['stock_actual'])

            messages.success(request, f"Importación completada: {actualizados} actualizado(s), {ignorados} ignorado(s).")
            return redirect(f"{reverse('inventario:lista_refacciones')}?area={area.pk}")

        except Exception as e:
            messages.error(request, f"Error al leer el archivo: {e}")
            return redirect(f"{reverse('inventario:importar_stock')}?area={area_id}")

    ctx = {
        'areas': Area.objects.filter(activa=True),
        'area_id': area_id,
    }
    return render(request, 'inventario_app/importar_stock.html', ctx)


@login_required
def descargar_plantilla_stock(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"

    headers = ['No. Item', 'Nombre', 'Stock actual']
    ws.append(headers)
    ws.append(['001', 'Rodamiento 6205', 12])

    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='4F46E5')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    anchos = [16, 36, 16]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Plantilla_Stock.xlsx"'
    return response


_UNIDAD_LABEL_A_CODIGO = {str(label).strip().lower(): codigo for codigo, label in Refaccion.UNIDAD_CHOICES}
_UNIDAD_CODIGOS = {codigo for codigo, label in Refaccion.UNIDAD_CHOICES}


@login_required
def importar_completo(request):
    acceso = getattr(request.user, 'acceso_mto', None)
    puede_editar = (
        request.user.is_superuser or
        (hasattr(request.user, 'perfil') and request.user.perfil.es_admin) or
        (acceso and acceso.editar_inventario)
    )
    if not puede_editar:
        messages.error(request, "No tienes permiso para importar inventario.")
        return redirect('inventario:lista_refacciones')

    area_id = request.GET.get('area', '') or request.POST.get('area', '')

    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        area_id = request.POST.get('area', '').strip()

        if not archivo or not area_id:
            messages.error(request, "Debes seleccionar un área y subir un archivo.")
            return redirect(f"{reverse('inventario:importar_completo')}?area={area_id}")

        try:
            area = get_object_or_404(Area, pk=area_id)
            wb = load_workbook(archivo, data_only=True)
            ws = wb.active

            categorias_dict = {}
            for c in CategoriaRefaccion.objects.all():
                for nombre_variante in (getattr(c, 'nombre_es', None), getattr(c, 'nombre_en', None), c.nombre):
                    if nombre_variante:
                        categorias_dict.setdefault(str(nombre_variante).strip().lower(), c)

            # Una sola lectura de las refacciones existentes del área — evita 1 query por fila.
            refacciones_existentes = {r.no_item: r for r in Refaccion.objects.filter(area=area)}
            nuevos_dict = {}
            ids_actualizados = set()
            errores = []

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row is None or not any(c not in (None, '') for c in row):
                    continue

                no_item = str(row[0]).strip() if len(row) > 0 and row[0] not in (None, '') else ''
                if not no_item:
                    continue

                nombre = str(row[1]).strip() if len(row) > 1 and row[1] not in (None, '') else ''
                if not nombre:
                    errores.append(f"Fila {i}: falta el nombre, se omitió.")
                    continue

                descripcion = str(row[2]).strip() if len(row) > 2 and row[2] not in (None, '') else ''

                categoria = None
                cat_txt = str(row[3]).strip() if len(row) > 3 and row[3] not in (None, '') else ''
                if cat_txt:
                    categoria = categorias_dict.get(cat_txt.lower())
                    if categoria is None:
                        errores.append(f"Fila {i}: categoría '{cat_txt}' no encontrada, se dejó sin categoría.")

                unidad_txt = str(row[4]).strip() if len(row) > 4 and row[4] not in (None, '') else ''
                unidad = 'pza'
                if unidad_txt:
                    low = unidad_txt.lower()
                    if low in _UNIDAD_CODIGOS:
                        unidad = low
                    elif low in _UNIDAD_LABEL_A_CODIGO:
                        unidad = _UNIDAD_LABEL_A_CODIGO[low]
                    else:
                        errores.append(f"Fila {i}: unidad '{unidad_txt}' no reconocida, se usó 'Pieza'.")

                try:
                    stock_actual = int(float(row[5])) if len(row) > 5 and row[5] not in (None, '') else 0
                    stock_minimo = int(float(row[6])) if len(row) > 6 and row[6] not in (None, '') else 0
                    stock_maximo = int(float(row[7])) if len(row) > 7 and row[7] not in (None, '') else 0
                except (ValueError, TypeError):
                    errores.append(f"Fila {i}: stock inválido, se omitió.")
                    continue

                ubicacion = str(row[8]).strip() if len(row) > 8 and row[8] not in (None, '') else ''
                if len(ubicacion) > 4:
                    errores.append(f"Fila {i}: la ubicación '{ubicacion}' supera 4 caracteres, se recortó.")
                    ubicacion = ubicacion[:4]

                proveedor = str(row[9]).strip() if len(row) > 9 and row[9] not in (None, '') else ''
                if len(proveedor) > 50:
                    proveedor = proveedor[:50]

                costo_unitario = None
                if len(row) > 10 and row[10] not in (None, ''):
                    try:
                        costo_unitario = float(row[10])
                    except (ValueError, TypeError):
                        errores.append(f"Fila {i}: costo unitario inválido, se dejó vacío.")

                if no_item in refacciones_existentes:
                    obj = refacciones_existentes[no_item]
                    ids_actualizados.add(no_item)
                elif no_item in nuevos_dict:
                    obj = nuevos_dict[no_item]
                else:
                    obj = Refaccion(no_item=no_item, area=area)
                    nuevos_dict[no_item] = obj

                obj.nombre         = nombre
                obj.descripcion    = descripcion
                obj.categoria      = categoria
                obj.unidad         = unidad
                obj.stock_actual   = stock_actual
                obj.stock_minimo   = stock_minimo
                obj.stock_maximo   = stock_maximo
                obj.ubicacion      = ubicacion
                obj.proveedor      = proveedor
                obj.costo_unitario = costo_unitario

            with transaction.atomic():
                if ids_actualizados:
                    Refaccion.objects.bulk_update(
                        [refacciones_existentes[n] for n in ids_actualizados],
                        ['nombre', 'descripcion', 'categoria', 'unidad',
                         'stock_actual', 'stock_minimo', 'stock_maximo',
                         'ubicacion', 'proveedor', 'costo_unitario']
                    )
                if nuevos_dict:
                    Refaccion.objects.bulk_create(list(nuevos_dict.values()))

            creados      = len(nuevos_dict)
            actualizados = len(ids_actualizados)

            resumen = f"Importación completada: {creados} creada(s), {actualizados} actualizada(s)."
            if errores:
                resumen += f" {len(errores)} advertencia(s)."
            messages.success(request, resumen)
            for e in errores[:15]:
                messages.warning(request, e)

            return redirect(f"{reverse('inventario:lista_refacciones')}?area={area.pk}")

        except Exception as e:
            messages.error(request, f"Error al leer el archivo: {e}")
            return redirect(f"{reverse('inventario:importar_completo')}?area={area_id}")

    ctx = {
        'areas': Area.objects.filter(activa=True),
        'area_id': area_id,
    }
    return render(request, 'inventario_app/importar_completo.html', ctx)


@login_required
def descargar_plantilla_completa(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    headers = ['No. Item', 'Nombre', 'Descripción', 'Categoría', 'Unidad',
               'Stock actual', 'Stock mínimo', 'Stock máximo', 'Ubicación', 'Proveedor', 'Costo unitario']
    ws.append(headers)
    ws.append(['001', 'Rodamiento 6205', 'Rodamiento de bolas blindado', 'Rodamientos', 'Pieza', 12, 5, 50, 'A12', 'Proveedor S.A.', 45.50])

    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='4F46E5')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    anchos = [12, 30, 36, 20, 12, 12, 12, 12, 12, 20, 14]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Plantilla_Inventario_Completo.xlsx"'
    return response


@login_required
def exportar_refacciones(request):
    acceso = getattr(request.user, 'acceso_mto', None)
    puede_ver = (
        request.user.is_superuser or
        (hasattr(request.user, 'perfil') and request.user.perfil.es_admin) or
        (acceso and acceso.ver_inventario)
    )
    if not puede_ver:
        messages.error(request, "No tienes permiso para exportar.")
        return redirect('inventario:lista_refacciones')

    area_id       = request.GET.get('area', '')
    categoria_id  = request.GET.get('categoria', '')
    busqueda      = request.GET.get('q', '').strip()
    estatus_stock = request.GET.get('estatus_stock', '')

    refacciones_qs = Refaccion.objects.select_related('area', 'categoria').filter(activo=True)
    if area_id:
        refacciones_qs = refacciones_qs.filter(area_id=area_id)
    if categoria_id:
        refacciones_qs = refacciones_qs.filter(categoria_id=categoria_id)
    if busqueda:
        refacciones_qs = (refacciones_qs.filter(nombre__icontains=busqueda)  |
                          refacciones_qs.filter(no_item__icontains=busqueda) |
                          refacciones_qs.filter(descripcion__icontains=busqueda))

    refacciones_lista = list(refacciones_qs)
    if estatus_stock == 'bajo_minimo':
        refacciones_lista = [r for r in refacciones_lista if r.bajo_minimo]
    elif estatus_stock == 'sobre_maximo':
        refacciones_lista = [r for r in refacciones_lista if r.sobre_maximo]
    elif estatus_stock == 'en_rango':
        refacciones_lista = [r for r in refacciones_lista if not r.bajo_minimo and not r.sobre_maximo and r.stock_actual > 0]
    elif estatus_stock == 'sin_stock':
        refacciones_lista = [r for r in refacciones_lista if r.stock_actual == 0]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventario'
    cabeceras = ['No. Item', 'Nombre', 'Descripción', 'Categoría', 'Área', 'Unidad',
                 'Stock actual', 'Stock mínimo', 'Stock máximo', 'Ubicación', 'Proveedor', 'Costo unitario']
    _estilo_cabecera(ws, cabeceras, [12, 30, 40, 20, 18, 10, 12, 12, 12, 12, 20, 14])

    for r in refacciones_lista:
        ws.append([
            r.no_item,
            r.nombre,
            r.descripcion,
            r.categoria.nombre if r.categoria else '',
            r.area.nombre,
            r.get_unidad_display(),
            r.stock_actual,
            r.stock_minimo,
            r.stock_maximo,
            r.ubicacion,
            r.proveedor,
            float(r.costo_unitario) if r.costo_unitario is not None else '',
        ])

    response = _excel_response('Inventario.xlsx')
    wb.save(response)
    return response