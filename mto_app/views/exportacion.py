from datetime import date, timedelta
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
import os
import io
import uuid
import threading
import tempfile

from ..models import PlanMantenimiento, RegistroEjecucion, PasoRutina, Responsable
from .utils import lunes_de_semana, _planes_que_tocan

# token → {'status': 'pending'|'ready'|'error', 'path': str, 'filename': str, 'error': str}
_jobs: dict = {}

MESES_ES = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
    5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
    9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
}


def _run_exportar(token, planes_semana, semana, anio, lunes, creador, static_dir):
    """Genera el Excel en background y guarda el resultado en _jobs."""
    try:
        domingo    = lunes + timedelta(days=6)
        ultimo_dia = domingo.day

        if lunes.month == domingo.month:
            texto_semana = (
                f"Semana: {semana} del {lunes.day} al {domingo.day} "
                f"de {MESES_ES[domingo.month]}"
            )
        else:
            texto_semana = (
                f"S{semana} del {lunes.day} de {MESES_ES[lunes.month]} "
                f"al {domingo.day} de {MESES_ES[domingo.month]}"
            )

        ids_planes = [p.id for p in planes_semana]
        areas_ids  = list(set(p.area_id for p in planes_semana))
        pts        = list(set(p.plan_trabajo for p in planes_semana))

        registros_map = {
            r.plan_id: r
            for r in RegistroEjecucion.objects.filter(
                plan_id__in=ids_planes, semana_inicio=lunes
            ).select_related('responsable')
        }

        pasos_map = {}
        for paso in PasoRutina.objects.filter(
            area_id__in=areas_ids, plan_trabajo__in=pts
        ).order_by('secuencia'):
            clave = (paso.area_id, paso.plan_trabajo)
            pasos_map.setdefault(clave, []).append(paso)

        supervisores_map = {}
        for sup in Responsable.objects.filter(
            area_id__in=list(set(p.area_id for p in planes_semana)),
            posicion__icontains='supervisor',
            activo=True
        ).order_by('area_id', 'apellidos'):
            if sup.area_id not in supervisores_map:
                supervisores_map[sup.area_id] = f"{sup.nombre} {sup.apellidos}"

        filas = []
        for plan in planes_semana:
            reg = registros_map.get(plan.id)
            responsable_ejecutor = ''
            if reg and reg.responsable:
                responsable_ejecutor = f"{reg.responsable.nombre} {reg.responsable.apellidos}"
            filas.append({
                'plan':        plan,
                'codigo':      plan.codigo,
                'actividad':   plan.actividad or '',
                'responsable': responsable_ejecutor,
                'supervisor':  supervisores_map.get(plan.area_id, ''),
                'registro':    reg,
            })

        plantilla_path = os.path.join(static_dir, 'Week Format.xlsm')
        wb = load_workbook(plantilla_path, keep_vba=True)

        px = lambda n: n * 9525

        def _load_img(nombre):
            p = os.path.join(static_dir, nombre)
            return open(p, 'rb').read() if os.path.exists(p) else None

        logo_bytes    = _load_img('logo.png')
        version_bytes = _load_img('Portada_Version.png')
        titulo_bytes  = _load_img('Portada_Titulo_Plan_Mantenimiento.png')

        MAX_POR_PORTADA_P1 = 42
        MAX_POR_PORTADA_P2 = 84
        FILA_INICIO_P1     = 10
        FILA_INICIO_P2     = 94
        PASO_FILA_P        = 2

        for portada_idx, nombre_hoja in enumerate(['Portada 1', 'Portada 2']):
            if nombre_hoja not in wb.sheetnames:
                continue

            max_portada = MAX_POR_PORTADA_P1 if portada_idx == 0 else MAX_POR_PORTADA_P2
            ws     = wb[nombre_hoja]
            inicio = portada_idx * MAX_POR_PORTADA_P1
            bloque = filas[inicio: inicio + max_portada]

            if not bloque:
                ws.sheet_state = 'hidden'
                continue

            ws.sheet_state = 'visible'
            ws['J2'] = ultimo_dia
            ws['E8'] = texto_semana

            if logo_bytes:
                img        = Image(io.BytesIO(logo_bytes))
                img.anchor = AbsoluteAnchor(
                    pos=XDRPoint2D(x=px(1200), y=px(0)),
                    ext=XDRPositiveSize2D(cx=px(120), cy=px(77))
                )
                ws.add_image(img)
            if version_bytes:
                img        = Image(io.BytesIO(version_bytes))
                img.anchor = AbsoluteAnchor(
                    pos=XDRPoint2D(x=px(10), y=px(10)),
                    ext=XDRPositiveSize2D(cx=px(100), cy=px(55))
                )
                ws.add_image(img)
            if titulo_bytes:
                img        = Image(io.BytesIO(titulo_bytes))
                img.anchor = AbsoluteAnchor(
                    pos=XDRPoint2D(x=px(350), y=px(5)),
                    ext=XDRPositiveSize2D(cx=px(633), cy=px(65))
                )
                ws.add_image(img)

            fila_inicio_portada = FILA_INICIO_P1 if portada_idx == 0 else FILA_INICIO_P2

            for col in range(1, 9):
                celda = ws.cell(row=4, column=col)
                celda.border = Border(
                    bottom=Side(style='medium', color='000000'),
                    top=celda.border.top   if celda.border else Side(),
                    left=celda.border.left if celda.border else Side(),
                    right=celda.border.right if celda.border else Side(),
                )

            if portada_idx == 1:
                for fila in range(70, 140):
                    ws.row_dimensions[fila].hidden = True

            for i in range(max_portada):
                fila_p = fila_inicio_portada + i * PASO_FILA_P
                fila_s = fila_p + 1
                if i < len(bloque):
                    entrada = bloque[i]
                    ws[f'B{fila_p}'] = entrada['codigo']
                    ws[f'C{fila_p}'] = entrada['actividad']
                    ws[f'E{fila_p}'] = entrada['responsable']
                    ws[f'D{fila_p}'] = entrada['supervisor']
                    ws.row_dimensions[fila_p].hidden = False
                    ws.row_dimensions[fila_s].hidden = False
                else:
                    ws[f'B{fila_p}'] = ''
                    ws[f'C{fila_p}'] = ''
                    ws[f'D{fila_p}'] = ''
                    ws[f'E{fila_p}'] = ''
                    ws.row_dimensions[fila_p].hidden = True
                    ws.row_dimensions[fila_s].hidden = True

            ws.print_area = 'A1:I93' if portada_idx == 0 else 'A1:I177'
            ws.page_setup.paperSize   = ws.PAPERSIZE_LETTER
            ws.page_setup.orientation = 'portrait'
            ws.page_setup.fitToWidth  = 1
            ws.page_setup.fitToHeight = 1
            ws.page_setup.scale       = 100
            if ws.sheet_properties is None:
                ws.sheet_properties = WorksheetProperties()
            if ws.sheet_properties.pageSetUpPr is None:
                ws.sheet_properties.pageSetUpPr = PageSetupProperties()
            ws.sheet_properties.pageSetUpPr.fitToPage = True

        hoja_base_nombre = next(
            (ws.title for ws in wb.worksheets if ws.title not in ('Portada 1', 'Portada 2')),
            None
        )
        if not hoja_base_nombre:
            _jobs[token] = {'status': 'error', 'error': 'La plantilla no tiene hojas de órdenes.'}
            return

        hoja_base = wb[hoja_base_nombre]

        hojas_existentes = [ws for ws in wb.worksheets if ws.title not in ('Portada 1', 'Portada 2')]
        for ws in hojas_existentes:
            ws.sheet_state = 'hidden'

        hojas_ordenes = []
        for idx in range(1, len(filas) + 1):
            nueva_hoja = wb.copy_worksheet(hoja_base)
            nueva_hoja.title = f'OT {idx}'
            nueva_hoja.sheet_state = 'visible'
            hojas_ordenes.append(nueva_hoja)

        fila_inicio_o = 16
        max_pasos     = 19

        for idx, entrada in enumerate(filas, start=1):
            plan     = entrada['plan']
            registro = entrada['registro']
            ws       = hojas_ordenes[idx - 1]

            for row_dim in ws.row_dimensions.values():
                row_dim.hidden = False
            for col_dim in ws.column_dimensions.values():
                col_dim.hidden = False

            duracion_h = round(plan.duracion_minutos / 60, 1)
            mes        = lunes.strftime('%m')
            aa         = lunes.strftime('%y')
            no_orden   = f"{semana:02d}{mes}{aa}{100 + idx}"

            ws['D6']  = semana
            ws['D7']  = no_orden
            ws['H7']  = creador
            ws['D8']  = plan.area.nombre
            ws['H8']  = '________________'
            ws['D9']  = plan.estatus
            ws['H9']  = duracion_h
            ws['D10'] = plan.prioridad
            ws['H10'] = plan.locacion or plan.area.nombre
            ws['D11'] = plan.tipo_mto
            ws['D12'] = plan.rutina
            ws['D13'] = plan.actividad
            ws['G13'] = plan.codigo
            ws['H13'] = plan.nombre_equipo or plan.actividad

            if logo_bytes:
                img        = Image(io.BytesIO(logo_bytes))
                img.anchor = AbsoluteAnchor(
                    pos=XDRPoint2D(x=px(940), y=px(0)),
                    ext=XDRPositiveSize2D(cx=px(120), cy=px(76))
                )
                ws.add_image(img)

            for col in range(3, 9):
                celda = ws.cell(row=4, column=col)
                celda.border = Border(
                    bottom=Side(style='medium', color='000000'),
                    top=celda.border.top   if celda.border else Side(),
                    left=celda.border.left if celda.border else Side(),
                    right=celda.border.right if celda.border else Side(),
                )

            if registro and registro.responsable:
                ws['C41'] = registro.responsable.numero_nomina
                ws['D41'] = f"{registro.responsable.nombre} {registro.responsable.apellidos}"
            else:
                ws['C41'] = '____________'
                ws['D41'] = '___________________________'

            pasos = pasos_map.get((plan.area_id, plan.plan_trabajo), [])

            for i in range(max_pasos):
                fila = fila_inicio_o + i
                ws[f'C{fila}'] = ''
                ws[f'D{fila}'] = ''
                ws[f'E{fila}'] = ''
                ws[f'F{fila}'] = '▭'
                ws[f'G{fila}'] = ''
                ws[f'H{fila}'] = ''
                ws[f'C{fila}'].alignment = Alignment(vertical='center', horizontal='center')
                ws[f'D{fila}'].alignment = Alignment(wrap_text=True, vertical='center')
                ws[f'F{fila}'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'F{fila}'].font      = Font(size=36)
                ws[f'H{fila}'].alignment = Alignment(wrap_text=True, vertical='center')
                ws.row_dimensions[fila].height = 25
                ws.row_dimensions[fila].hidden = False

            for i, paso in enumerate(pasos[:max_pasos]):
                fila = fila_inicio_o + i
                ws[f'C{fila}'] = paso.secuencia
                ws[f'D{fila}'] = paso.descripcion
                ws[f'F{fila}'] = '▭'
                ws[f'G{fila}'] = ''
                ws[f'H{fila}'] = paso.detalles
                ws[f'C{fila}'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                ws[f'D{fila}'].alignment = Alignment(wrap_text=True, vertical='center')
                ws[f'F{fila}'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'F{fila}'].font      = Font(size=36)
                ws[f'H{fila}'].alignment = Alignment(wrap_text=True, vertical='center')
                desc    = paso.descripcion or ''
                detalle = paso.detalles or ''
                lineas  = max(
                    max(1, -(-len(desc) // 35)),
                    max(1, -(-len(detalle) // 85) + detalle.count('\n'))
                )
                ws.row_dimensions[fila].height = max(25, lineas * 14)

            pasos_count = min(len(pasos), max_pasos)
            for i in range(pasos_count, max_pasos):
                ws.row_dimensions[fila_inicio_o + i].hidden = True

            ws.print_area = 'A1:I52'
            ws.page_setup.paperSize   = ws.PAPERSIZE_LETTER
            ws.page_setup.orientation = 'portrait'
            ws.page_setup.fitToWidth  = 1
            ws.page_setup.fitToHeight = 0
            ws.page_setup.scale       = 100

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsm')
        wb.save(tmp)
        tmp.close()

        _jobs[token] = {
            'status':   'ready',
            'path':     tmp.name,
            'filename': f"OTs_Semana_{semana}_{anio}.xlsm",
        }

    except Exception as e:
        _jobs[token] = {'status': 'error', 'error': str(e)}


@login_required
def exportar_semana_excel(request):
    semana  = int(request.GET.get('semana', date.today().isocalendar()[1]))
    anio    = int(request.GET.get('anio',   date.today().isocalendar()[0]))
    area_id = request.GET.get('area')

    lunes         = lunes_de_semana(anio, semana)
    planes_qs     = PlanMantenimiento.objects.select_related('area').filter(activo=True)
    if area_id:
        planes_qs = planes_qs.filter(area_id=area_id)
    planes_semana = _planes_que_tocan(list(planes_qs), lunes)

    if not planes_semana:
        return JsonResponse(
            {'error': f'No hay planes para la semana {semana}/{anio}.'},
            status=404
        )

    creador    = request.user.get_full_name() or request.user.username
    static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'static', 'mto_app')
    token      = str(uuid.uuid4())
    _jobs[token] = {'status': 'pending'}

    threading.Thread(
        target=_run_exportar,
        args=(token, planes_semana, semana, anio, lunes, creador, static_dir),
        daemon=True,
    ).start()

    return JsonResponse({'token': token})


@login_required
def exportar_semana_status(request, token):
    job = _jobs.get(token, {'status': 'error', 'error': 'Token no encontrado'})
    return JsonResponse({'status': job['status'], 'error': job.get('error', '')})


@login_required
def exportar_semana_download(request, token):
    job = _jobs.pop(token, None)
    if not job or job['status'] != 'ready':
        messages.error(request, 'Archivo no disponible o aún generándose.')
        return redirect('mto:lista_plan')

    path     = job['path']
    filename = job['filename']
    with open(path, 'rb') as f:
        content = f.read()
    os.unlink(path)

    response = HttpResponse(
        content,
        content_type='application/vnd.ms-excel.sheet.macroEnabled.12'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
