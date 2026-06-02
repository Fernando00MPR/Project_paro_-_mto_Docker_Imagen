from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from openpyxl import load_workbook
from .utils import areas_permitidas_mto
from ..models import Area, PlanMantenimiento, PasoRutina, HistorialRutina

@login_required
def lista_pasos(request):
    plan_trabajo  = request.GET.get('pt', '').strip()
    area_id       = request.GET.get('area', '').strip()
    filtro_codigo = request.GET.get('codigo', '').strip()
    filtro_nombre = request.GET.get('nombre', '').strip()

    if area_id:
        request.session['pasos_area_id'] = area_id
    else:
        area_id = request.session.get('pasos_area_id', '')

    pasos = PasoRutina.objects.select_related('area').none()
    if plan_trabajo:
        pasos = PasoRutina.objects.select_related('area').filter(plan_trabajo=plan_trabajo)
        if area_id:
            pasos = pasos.filter(area_id=area_id)

    from django.db.models import Max

    planes_ids = PasoRutina.objects.filter(
        **({'area_id': area_id} if area_id else {})
    ).values('plan_trabajo').annotate(id=Max('id')).values_list('id', flat=True)

    planes_qs = PasoRutina.objects.filter(
        id__in=planes_ids
    ).values('plan_trabajo', 'codigo_plan', 'nombre_plan')

    # Filtros de código y nombre — DESPUÉS de construir planes_qs
    if filtro_codigo:
        planes_qs = planes_qs.filter(codigo_plan__icontains=filtro_codigo)
    if filtro_nombre:
        planes_qs = planes_qs.filter(nombre_plan__icontains=filtro_nombre)

    planes_con_pasos = sorted(
        planes_qs,
        key=lambda x: int(x['plan_trabajo']) if str(x['plan_trabajo']).isdigit() else x['plan_trabajo']
    )

    per_page_raw = request.GET.get('per_page', '10')
    per_page     = per_page_raw if per_page_raw in ('10', '15', '20') else '10'
    paginator    = Paginator(pasos, int(per_page))
    page_num     = request.GET.get('page', 1)
    pasos_page   = paginator.get_page(page_num)

    primer_paso = pasos.first() if plan_trabajo else None

    h_plan    = request.GET.get('h_plan', '').strip()
    h_codigo  = request.GET.get('h_codigo', '').strip()
    h_nombre  = request.GET.get('h_nombre', '').strip()

    historial_qs = HistorialRutina.objects.select_related('usuario', 'area').all()
    if h_plan:
        historial_qs = historial_qs.filter(plan_trabajo__icontains=h_plan)
    if h_codigo:
        historial_qs = historial_qs.filter(codigo_plan__icontains=h_codigo)
    if h_nombre:
        historial_qs = historial_qs.filter(nombre_plan__icontains=h_nombre)

    h_per_page_raw = request.GET.get('h_per_page', '5')
    h_per_page     = h_per_page_raw if h_per_page_raw in ('5', '10', '15', '20') else '5'
    h_paginator    = Paginator(historial_qs, int(h_per_page))
    h_page_num     = request.GET.get('h_page', 1)
    historial_page = h_paginator.get_page(h_page_num)

    ctx = {
        'pasos':            pasos_page,
        'planes_con_pasos': planes_con_pasos,
        'areas':            areas_permitidas_mto(request),
        'filtro_pt':        plan_trabajo,
        'filtro_area':      area_id,
        'per_page':         per_page,
        'perpage_opts':     ['10', '15', '20'],
        'primer_paso':      primer_paso,
        'filtro_codigo':    filtro_codigo,
        'filtro_nombre':    filtro_nombre,
        'historial':        historial_page,
        'h_plan':           h_plan,
        'h_codigo':         h_codigo,
        'h_nombre':         h_nombre,
        'h_per_page':       h_per_page,
        'h_perpage_opts':   ['5', '10', '15', '20'],
    }
    return render(request, 'mto_app/rutinas/lista_rutinas.html', ctx)


@login_required
def importar_pasos(request):

    def limpiar(valor):
        if not valor:
            return ''
        #return str(valor).replace('_x000D_', '').replace('_x000d_', '').replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ').strip()
        return str(valor).replace('_x000D_', '').replace('_x000d_', '').strip()

    if request.method == 'POST':
        archivo      = request.FILES.get('archivo')
        plan_trabajo = request.POST.get('plan_trabajo', '').strip()
        area_id      = request.POST.get('area', '').strip()
        reemplazar   = 'reemplazar' in request.POST

        if not archivo or not plan_trabajo or not area_id:
            messages.error(request, "Debes indicar el area, N° plan y subir un archivo.")
            return redirect('mto:importar_pasos')

        try:
            area = Area.objects.get(pk=area_id)
            wb   = load_workbook(archivo, data_only=True)
            ws   = wb.active

            nombre_plan = limpiar(ws.cell(row=5, column=4).value)
            codigo_plan = limpiar(ws.cell(row=6, column=4).value)

            if reemplazar:
                PasoRutina.objects.filter(area=area, plan_trabajo=plan_trabajo).delete()

            creados = 0
            for row in ws.iter_rows(min_row=11, values_only=True):
                seq      = row[1]
                desc     = row[2]
                detalles = row[6] if len(row) > 6 and row[6] else ''

                if not seq or not desc:
                    continue
                try:
                    seq = int(seq)
                except Exception:
                    continue

                PasoRutina.objects.update_or_create(
                    area=area,
                    plan_trabajo=plan_trabajo,
                    secuencia=seq,
                    defaults={
                        'descripcion': limpiar(desc),
                        'detalles':    limpiar(detalles),
                        'nombre_plan': nombre_plan,
                        'codigo_plan': codigo_plan,
                    }
                )
                creados += 1

            messages.success(request,
                f"{creados} paso(s) importados para el plan {plan_trabajo} - {area.nombre}.")

            motivo = request.POST.get('motivo', '').strip()
            HistorialRutina.objects.create(
                usuario=request.user,
                accion='importar',
                area=area,
                plan_trabajo=plan_trabajo,
                codigo_plan=codigo_plan,
                nombre_plan=nombre_plan,
                pasos_afectados=creados,
                motivo=motivo,
            )

            return redirect('mto:lista_pasos')

        except Exception as e:
            messages.error(request, f"Error al leer el archivo: {e}")

    planes = sorted(
        set(PlanMantenimiento.objects.values_list('plan_trabajo', flat=True)),
        key=lambda x: int(x) if str(x).isdigit() else x
    )
    areas = areas_permitidas_mto(request)
    return render(request, 'mto_app/rutinas/importar_rutinas.html', {
        'planes': planes,
        'areas':  areas,
    })


@login_required
def eliminar_pasos_plan(request):
    if request.method == 'POST':
        plan_trabajo = request.POST.get('plan_trabajo', '').strip()
        area_id      = request.POST.get('area_id', '').strip()
        motivo       = request.POST.get('motivo', '').strip()
        if plan_trabajo and area_id:
            pasos_qs = PasoRutina.objects.filter(
                plan_trabajo=plan_trabajo, area_id=area_id
            )
            primer = pasos_qs.first()
            codigo_plan = primer.codigo_plan if primer else ''
            nombre_plan = primer.nombre_plan if primer else ''
            try:
                area = Area.objects.get(pk=area_id)
            except Area.DoesNotExist:
                area = None

            cantidad = pasos_qs.count()
            pasos_qs.delete()

            HistorialRutina.objects.create(
                usuario=request.user,
                accion='eliminar',
                area=area,
                plan_trabajo=plan_trabajo,
                codigo_plan=codigo_plan,
                nombre_plan=nombre_plan,
                pasos_afectados=cantidad,
                motivo=motivo,
            )

            messages.success(request, f"{cantidad} paso(s) eliminados.")
    return redirect('mto:lista_pasos')