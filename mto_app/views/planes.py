import io
from datetime import date, timedelta
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.views.decorators.clickjacking import xframe_options_exempt
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from collections import defaultdict
from .utils import areas_permitidas_mto
from ..models import Area, PlanMantenimiento, RegistroEjecucion, Responsable
from django.urls import reverse
from .utils import lunes_de_semana, INTERVALO, _frecuencia_desde_excel, _aplicar_orden_plan, _orden_por_plan_trabajo


@login_required
def lista_plan(request):
    acceso = getattr(request.user, 'acceso_mto', None)
    puede_ver = (
        request.user.is_superuser or
        (hasattr(request.user, 'perfil') and request.user.perfil.es_admin) or
        (acceso and acceso.ver_plan_mantenimiento)
    )
    if not puede_ver:
        messages.error(request, "No tienes permiso para ver esta sección.")
        return redirect('mto:dashboard')
    
    area_id       = request.GET.get('area')
    frecuencia    = request.GET.get('frecuencia')
    busqueda      = request.GET.get('q', '').strip()
    filtro_semana = request.GET.get('semana', '').strip()
    filtro_anio   = request.GET.get('anio', '').strip()
    orden         = request.GET.get('orden', '').strip()

    planes = PlanMantenimiento.objects.select_related('area').filter(activo=True)

    if area_id:
        request.session['lista_plan_area_id'] = area_id
    else:
        area_id = request.session.get('lista_plan_area_id', '')

    if area_id:
        planes = planes.filter(area_id=area_id)
    if frecuencia:
        planes = planes.filter(frecuencia=frecuencia)
    if busqueda:
        planes = (planes.filter(actividad__icontains=busqueda) |
                planes.filter(codigo__icontains=busqueda) |
                planes.filter(rutina__icontains=busqueda))

    planes = _aplicar_orden_plan(planes, request.GET)
    
    hoy         = date.today()
    anio_filtro = int(filtro_anio) if filtro_anio else hoy.year

    planes_list = list(planes)
    areas_mto   = list(areas_permitidas_mto(request))

    responsables_por_area_map  = defaultdict(list)
    responsables_por_area_dict = defaultdict(list)
    for r in Responsable.objects.filter(activo=True).order_by('apellidos', 'nombre'):
        responsables_por_area_map[r.area_id].append(r)
        responsables_por_area_dict[r.area_id].append({
            'pk': r.pk, 'nombre': r.nombre, 'apellidos': r.apellidos,
            'numero_nomina': r.numero_nomina,
        })

    lunes_busq_pre       = None
    registros_semana_pre = {}
    if filtro_semana:
        try:
            lunes_busq_pre = lunes_de_semana(anio_filtro, int(filtro_semana))
            registros_semana_pre = {
                r.plan_id: r
                for r in RegistroEjecucion.objects.select_related('responsable').filter(
                    plan_id__in=[p.id for p in planes_list],
                    semana_inicio=lunes_busq_pre,
                )
            }
        except Exception:
            lunes_busq_pre = None

    planes_con_datos = []

    for plan in planes_list:
        plan.proximo               = None
        plan.semana_inicio_lunes   = None
        plan.semana_inicio_domingo = None
        plan.registro_semana       = None
        plan.semana_actual         = None
        plan.responsables_area     = responsables_por_area_map[plan.area_id]

        if plan.semana_inicio and plan.anio_inicio:
            try:
                intervalo = INTERVALO.get(plan.frecuencia, 1)
                ref_lunes = lunes_de_semana(plan.anio_inicio, plan.semana_inicio)

                plan.semana_inicio_lunes   = ref_lunes
                plan.semana_inicio_domingo = ref_lunes + timedelta(days=6)

                cursor = ref_lunes
                while cursor < hoy:
                    cursor += timedelta(weeks=intervalo)
                iso          = cursor.isocalendar()
                plan.proximo = {
                    'semana':  iso[1],
                    'lunes':   cursor,
                    'domingo': cursor + timedelta(days=6),
                }

                if filtro_semana:
                    if lunes_busq_pre is None:
                        continue
                    lunes_busq = lunes_busq_pre
                    delta   = (lunes_busq - ref_lunes).days
                    semanas = delta // 7
                    if delta < 0 or semanas % intervalo != 0:
                        continue

                    plan.semana_actual = {
                        'semana':  int(filtro_semana),
                        'lunes':   lunes_busq,
                        'domingo': lunes_busq + timedelta(days=6),
                    }

                    plan.registro_semana = registros_semana_pre.get(plan.id)

            except Exception:
                if filtro_semana:
                    continue

        elif filtro_semana:
            continue

        planes_con_datos.append(plan)

    stats_semana = None
    if filtro_semana and lunes_busq_pre:
            
            total         = len(planes_con_datos)
            completadas_s = sum(1 for p in planes_con_datos
                                if registros_semana_pre.get(p.pk) and registros_semana_pre[p.pk].estado == 'completada')
            pendientes_s  = total - completadas_s
            minutos_total = sum(p.duracion_minutos or 0 for p in planes_con_datos)
            stats_semana  = {
                'total':         total,
                'completadas':   completadas_s,
                'pendientes':    pendientes_s,
                'minutos_total': minutos_total,
                'horas':         minutos_total // 60,
                'minutos_resto': minutos_total % 60,
            }

    planes_con_datos = _orden_por_plan_trabajo(planes_con_datos, orden)

    per_page_raw = request.GET.get('per_page', '8')
    per_page     = per_page_raw if per_page_raw in ('8', '10', '15', '20') else '8'
    paginator    = Paginator(planes_con_datos, int(per_page))
    page_num     = request.GET.get('page', 1)
    planes_page  = paginator.get_page(page_num)

    ctx = {
        'planes':        planes_page,
        'anio_actual':   hoy.year,
        'areas':         areas_mto,
        'areas_menu':    areas_mto,
        'frecuencias':   PlanMantenimiento.FRECUENCIA_CHOICES,
        'filtro_area':   area_id,
        'filtro_frec':   frecuencia,
        'filtro_orden':  orden,
        'busqueda':      busqueda,
        'filtro_semana': filtro_semana,
        'filtro_anio':   anio_filtro,
        'stats_semana':  stats_semana,
        'per_page':      per_page,
        'perpage_opts':  ['8', '10', '15', '20'],
        'usuarios':      User.objects.filter(is_active=True).order_by('first_name'),
        'responsables_por_area': {
            str(a.pk): responsables_por_area_dict[a.id]
            for a in areas_mto
        },
        'puede_editar_plan': (
            request.user.is_superuser or
            (hasattr(request.user, 'perfil') and request.user.perfil.es_admin) or
            (acceso and acceso.editar_plan_mantenimiento)
        ),
        'puede_eliminar_plan': (
            request.user.is_superuser or
            (hasattr(request.user, 'perfil') and request.user.perfil.es_admin) or
            (acceso and acceso.eliminar_plan_mantenimiento)
        ),
        'puede_eliminar_todos_planes': (
            request.user.is_superuser or
            (hasattr(request.user, 'perfil') and request.user.perfil.es_admin)
        ),
        'total_planes_area': PlanMantenimiento.objects.filter(area_id=area_id).count() if area_id else 0,
    }
    return render(request, 'mto_app/plan/lista.html', ctx)


@login_required
def form_plan(request, pk=None):
    plan  = get_object_or_404(PlanMantenimiento, pk=pk) if pk else None
    areas = areas_permitidas_mto(request)

    if request.method == 'POST':
        try:
            semana_inicio = request.POST.get('semana_inicio', '').strip() or None
            anio_inicio   = request.POST.get('anio_inicio', '').strip() or None

            datos = {
                'area':             get_object_or_404(Area, pk=request.POST.get('area')),
                'codigo':           request.POST.get('codigo', '').strip(),
                'actividad':        request.POST.get('actividad', '').strip(),
                'rutina':           request.POST.get('rutina', '').strip(),
                'plan_trabajo':     request.POST.get('plan_trabajo', '').strip(),
                'frecuencia':       request.POST.get('frecuencia', ''),
                'duracion_minutos': int(request.POST.get('duracion_minutos', 60) or 60),
                'semana_inicio':    int(semana_inicio) if semana_inicio else None,
                'anio_inicio':      int(anio_inicio)   if anio_inicio   else None,
                'nombre_equipo':    request.POST.get('nombre_equipo', '').strip(),
                'locacion':         request.POST.get('locacion', '').strip(),
                'tipo_mto':         request.POST.get('tipo_mto', 'Preventivo'),
                'prioridad':        request.POST.get('prioridad', 'Baja'),
                'estatus':          request.POST.get('estatus', 'Abierta'),
                'activo':           'activo' in request.POST,
            }

            if not all([datos['codigo'], datos['actividad'], datos['rutina'], datos['frecuencia']]):
                raise ValueError("Todos los campos marcados son obligatorios.")

            if plan:
                for k, v in datos.items():
                    setattr(plan, k, v)
                plan.save()
                messages.success(request, f"Plan '{plan.codigo}' actualizado.")
            else:
                PlanMantenimiento.objects.create(**datos)
                messages.success(request, "Plan de mantenimiento creado.")
            return redirect('mto:lista_plan')

        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f"Error al guardar: {e}")

    ctx = {
        'plan':        plan,
        'areas':       areas,
        'frecuencias': PlanMantenimiento.FRECUENCIA_CHOICES,
        'tipos_mto':   PlanMantenimiento.TIPO_MTO_CHOICES,
        'prioridades': PlanMantenimiento.PRIORIDAD_CHOICES,
        'status_opts': PlanMantenimiento.STATUS_CHOICES,
    }
    return render(request, 'mto_app/plan/form.html', ctx)


@login_required
def eliminar_plan(request, pk):
    plan = get_object_or_404(PlanMantenimiento, pk=pk)
    if request.method == 'POST':
        try:
            plan.delete()
            messages.success(request, f"Plan '{plan.codigo}' eliminado.")
        except Exception:
            messages.error(request,
                f"No se puede eliminar '{plan.codigo}' porque tiene registros asociados.")
    return redirect('mto:lista_plan')


@login_required
def eliminar_todos_planes_area(request):
    es_admin_total = (
        request.user.is_superuser or
        (hasattr(request.user, 'perfil') and request.user.perfil.es_admin)
    )
    if not es_admin_total:
        messages.error(request, "No tienes permiso para eliminar todos los planes del área.")
        return redirect('mto:lista_plan')

    area_id = request.POST.get('area')
    if request.method == 'POST' and area_id:
        area = get_object_or_404(Area, pk=area_id)
        cantidad = PlanMantenimiento.objects.filter(area=area).count()
        PlanMantenimiento.objects.filter(area=area).delete()
        messages.success(
            request,
            f"Se eliminaron los {cantidad} planes de mantenimiento del área '{area.nombre}' y su historial de ejecuciones."
        )
    return redirect(f"{reverse('mto:lista_plan')}?area={area_id}")


@login_required
def exportar_plan(request):
    acceso = getattr(request.user, 'acceso_mto', None)
    puede_ver = (
        request.user.is_superuser or
        (hasattr(request.user, 'perfil') and request.user.perfil.es_admin) or
        (acceso and acceso.ver_plan_mantenimiento)
    )
    if not puede_ver:
        messages.error(request, "No tienes permiso para exportar planes de mantenimiento.")
        return redirect('mto:lista_plan')

    area_id = request.GET.get('area') or request.session.get('lista_plan_area_id', '')
    planes  = PlanMantenimiento.objects.select_related('area').filter(activo=True)
    if area_id:
        planes = planes.filter(area_id=area_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Plan de mantenimiento"

    # Mismas columnas que descargar_plantilla, en el mismo orden que lee importar_plan.
    headers = [
        '', 'CODIGO', 'RUTINA', 'EQUIPO (descripcion tarea)', 'Duracion (horas)',
        'Frecuencia', 'Plan de Trabajo', 'Equipo (nombre corto)', 'Locacion',
        'Tipo de mantenimiento', 'Prioridad', 'Status', 'Semana inicio', 'Anio inicio'
    ]
    ws.append([])  # Fila 1
    ws.append([])  # Fila 2
    ws.append(headers)  # Fila 3

    for plan in planes.order_by('codigo'):
        codigo_frecuencia = f"{INTERVALO.get(plan.frecuencia, 4)}S"
        ws.append([
            '',
            plan.codigo,
            plan.rutina,
            plan.actividad,
            round(plan.duracion_minutos / 60, 2),
            codigo_frecuencia,
            plan.plan_trabajo,
            plan.nombre_equipo,
            plan.locacion,
            plan.tipo_mto,
            plan.prioridad,
            plan.estatus,
            plan.semana_inicio or '',
            plan.anio_inicio or '',
        ])

    for cell in ws[3][1:]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='4F46E5')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[3].height = 30

    # Centrar todas las filas de datos (una por plan)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    anchos = [5, 14, 18, 40, 16, 12, 14, 30, 25, 22, 12, 12, 14, 12]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from django.http import HttpResponse
    nombre_area = Area.objects.filter(id=area_id).first() if area_id else None
    filename = f"Planes_{nombre_area.nombre.replace(' ','_')}.xlsx" if nombre_area else "Planes_Mantenimiento.xlsx"

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@xframe_options_exempt
def modal_plan(request, pk):
    plan = get_object_or_404(PlanMantenimiento, pk=pk)
    hoy  = date.today()
    anio = int(request.GET.get('anio', hoy.year))

    fechas = []
    if plan.semana_inicio and plan.anio_inicio:
        intervalo = INTERVALO.get(plan.frecuencia, 1)
        try:
            ref_lunes   = lunes_de_semana(plan.anio_inicio, plan.semana_inicio)
            inicio_anio = date(anio, 1, 1)
            fin_anio    = date(anio, 12, 31)
            registros_map = {
                r.semana_inicio: r
                for r in RegistroEjecucion.objects.filter(
                    plan=plan,
                    semana_inicio__gte=inicio_anio,
                    semana_inicio__lte=fin_anio,
                )
            }

            cursor      = ref_lunes
            while cursor <= fin_anio:
                if cursor.isocalendar()[0] == anio:

                    fechas.append({
                        'fecha':    cursor,
                        'domingo':  cursor + timedelta(days=6),
                        'semana':   cursor.isocalendar()[1],
                        'registro': registros_map.get(cursor),
                        'pasada':   cursor < hoy,
                        'actual':   (cursor.isocalendar()[1] == hoy.isocalendar()[1]
                                     and cursor.year == hoy.year),
                    })
                cursor += timedelta(weeks=intervalo)
        except Exception:
            pass

    proxima = next((f for f in fechas if not f['pasada']), None)

    ctx = {
        'plan':           plan,
        'fechas':         fechas,
        'proxima':        proxima,
        'anio':           anio,
        'anio_anterior':  anio - 1,
        'anio_siguiente': anio + 1,
    }
    return render(request, 'mto_app/plan/modal_calendario.html', ctx)


@login_required
def importar_plan(request):
    areas = areas_permitidas_mto(request)

    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        area_id = request.POST.get('area')

        if not archivo or not area_id:
            messages.error(request, "Debes seleccionar un archivo y un area.")
            return render(request, 'mto_app/plan/importar.html', {'areas': areas})

        try:
            area = Area.objects.get(pk=area_id)
            wb   = load_workbook(archivo, data_only=True)
            ws   = wb.active

            creados = actualizados = omitidos = 0
            errores = []

            codigos_existentes = {
                p.codigo: p.area.nombre
                for p in PlanMantenimiento.objects.select_related('area').exclude(area=area)
            }
            planes_del_area   = {p.codigo: p for p in PlanMantenimiento.objects.filter(area=area)}
            nuevos_por_codigo = {}
            a_actualizar_por_pk = {}
            campos_actualizables = [
                'area', 'actividad', 'rutina', 'duracion_minutos', 'frecuencia', 'plan_trabajo',
                'nombre_equipo', 'locacion', 'tipo_mto', 'prioridad', 'estatus', 'activo',
                'semana_inicio', 'anio_inicio',
            ]

            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row or not row[1]:
                    continue
                if str(row[1]).strip().upper() in ('CODIGO', '#', 'NONE', ''):
                    continue

                try:
                    codigo        = str(row[1]).strip()
                    rutina        = str(row[2]).strip()   if row[2]                     else ''
                    actividad     = str(row[3]).strip()   if row[3]                     else ''
                    duracion_raw  = row[4]
                    frecuencia_raw= row[5]
                    plan_trabajo  = str(row[6]).strip()   if len(row) > 6  and row[6]  else ''
                    nombre_equipo = str(row[7]).strip()   if len(row) > 7  and row[7]  else ''
                    locacion      = str(row[8]).strip()   if len(row) > 8  and row[8]  else ''
                    tipo_mto      = str(row[9]).strip()   if len(row) > 9  and row[9]  else 'Preventivo'
                    prioridad     = str(row[10]).strip()  if len(row) > 10 and row[10] else 'Baja'
                    estatus       = str(row[11]).strip()  if len(row) > 11 and row[11] else 'Abierta'

                    try:
                        semana_inicio = int(float(str(row[12]))) if len(row) > 12 and row[12] and str(row[12]).strip() not in ('', 'None') else None
                    except Exception:
                        semana_inicio = None
                    try:
                        anio_inicio = int(float(str(row[13]))) if len(row) > 13 and row[13] and str(row[13]).strip() not in ('', 'None') else None
                    except Exception:
                        anio_inicio = None

                    if not codigo or not actividad:
                        omitidos += 1
                        continue

                    if codigo in codigos_existentes:
                        errores.append(
                            f"Código '{codigo}' ya existe en el área '{codigos_existentes[codigo]}' "
                            f"— omitido para evitar sobreescritura."
                        )
                        omitidos += 1
                        continue

                    try:
                        duracion_h       = float(str(duracion_raw).replace(',', '.')) if duracion_raw else 1.0
                        duracion_minutos = max(5, int(duracion_h * 60))
                    except Exception:
                        duracion_minutos = 60

                    frecuencia = _frecuencia_desde_excel(frecuencia_raw)

                    if tipo_mto not in [t[0] for t in PlanMantenimiento.TIPO_MTO_CHOICES]:
                        tipo_mto = 'Preventivo'
                    if prioridad not in [p[0] for p in PlanMantenimiento.PRIORIDAD_CHOICES]:
                        prioridad = 'Baja'
                    if estatus not in [s[0] for s in PlanMantenimiento.STATUS_CHOICES]:
                        estatus = 'Abierta'

                    defaults = {
                        'area':             area,
                        'actividad':        actividad,
                        'rutina':           rutina,
                        'duracion_minutos': duracion_minutos,
                        'frecuencia':       frecuencia,
                        'plan_trabajo':     plan_trabajo,
                        'nombre_equipo':    nombre_equipo,
                        'locacion':         locacion,
                        'tipo_mto':         tipo_mto,
                        'prioridad':        prioridad,
                        'estatus':          estatus,
                        'activo':           True,
                    }
                    if semana_inicio:
                        defaults['semana_inicio'] = semana_inicio
                    if anio_inicio:
                        defaults['anio_inicio'] = anio_inicio

                    plan_existente = planes_del_area.get(codigo)
                    ya_pendiente    = nuevos_por_codigo.get(codigo)

                    if plan_existente:
                        for campo, valor in defaults.items():
                            setattr(plan_existente, campo, valor)
                        a_actualizar_por_pk[plan_existente.pk] = plan_existente
                        actualizados += 1
                    elif ya_pendiente:
                        # Mismo código repetido en el archivo — la última fila gana, sin contarlo dos veces.
                        for campo, valor in defaults.items():
                            setattr(ya_pendiente, campo, valor)
                    else:
                        nuevos_por_codigo[codigo] = PlanMantenimiento(codigo=codigo, **defaults)
                        creados += 1

                except Exception as e:
                    errores.append(f"Fila con codigo '{row[1]}': {e}")

            if nuevos_por_codigo:
                PlanMantenimiento.objects.bulk_create(nuevos_por_codigo.values())
            if a_actualizar_por_pk:
                PlanMantenimiento.objects.bulk_update(a_actualizar_por_pk.values(), campos_actualizables)

            resumen = f"{creados} creado(s), {actualizados} actualizado(s), {omitidos} omitido(s)."
            if errores:
                messages.warning(request, f"Importacion con advertencias — {resumen}")
                for err in errores[:5]:
                    messages.error(request, err)
            else:
                messages.success(request, f"Importacion exitosa — {resumen}")

            return redirect('mto:dashboard')

        except Exception as e:
            messages.error(request, f"Error al leer el archivo: {e}")

    return render(request, 'mto_app/plan/importar.html', {'areas': areas})


@login_required
def descargar_plantilla(request):
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan de mantenimiento"

    # Estas columnas deben coincidir exactamente, en orden, con los índices que
    # lee importar_plan() (row[1]=codigo, row[2]=rutina, ... row[13]=anio_inicio).
    headers = [
        '', 'CODIGO', 'RUTINA', 'EQUIPO (descripcion tarea)', 'Duracion (horas)',
        'Frecuencia', 'Plan de Trabajo', 'Equipo (nombre corto)', 'Locacion',
        'Tipo de mantenimiento', 'Prioridad', 'Status', 'Semana inicio', 'Anio inicio'
    ]

    ws.append([])  # Fila 1
    ws.append([])  # Fila 2
    ws.append(headers)
    ws.append([
        '', 'SWC-001', 'RMAS001-00P', 'Mantenimiento Mecanico a Transportador # 1',
        1.0, '16S', '1', 'Mesa Fija Conveyor # 01', 'Automatic Warehouse',
        'Preventivo', 'Baja', 'Abierta', 1, 2025
    ])

    ws.append([])
    ws.append(['','NOTA: Frecuencia: 1S=semanal, 2S=quincenal, 4S=mensual, 9S=bimestral, 13S=trimestral, 17S=cuatrimestral, 26S=semestral, 52S=anual'])

    for cell in ws[3][1:]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='4F46E5')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Centrar la fila de ejemplo también
    for cell in ws[4]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[3].height = 30

    anchos = [5, 14, 18, 40, 16, 12, 14, 30, 25, 22, 12, 12, 14, 12]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from django.http import HttpResponse
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Plantilla_Mantenimiento.xlsx"'
    return response


