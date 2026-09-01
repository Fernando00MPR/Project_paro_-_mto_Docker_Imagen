from datetime import date
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import render
from django.views.decorators.http import require_POST
import json

from ..models import Area, Paro, RegistroProduccion, CatalogoEquipo, CatalogoMolde, ConfiguracionMoldes
from login_app.permisos import get_perfil

RESPONSABLES_MANT = {'Mantenimiento', 'Maintenance', 'Robótica', 'Robotic', 'Robotics'}

def _area_moldes_id():
    config = ConfiguracionMoldes.objects.only('area_id').first()
    return config.area_id if config else None


def _normalizar_equipo(area, texto_equipo):
    """Busca el equipo en el catálogo (bilingüe) para guardar ambas
    traducciones en el registro, sin importar en qué idioma se envió el
    texto — mismo criterio que usa paros.py para Paro.equipo_es/equipo_en."""
    if not texto_equipo:
        return '', ''
    catalogo = CatalogoEquipo.objects.filter(area=area).filter(
        Q(equipo_es=texto_equipo) | Q(equipo_en=texto_equipo)
    ).first()
    if catalogo:
        return (
            catalogo.equipo_es or catalogo.equipo_en or texto_equipo,
            catalogo.equipo_en or catalogo.equipo_es or texto_equipo,
        )
    return texto_equipo, texto_equipo


def _minutos(t):
    return t.hour * 60 + t.minute


def _rango_minutos(ini, fin):
    a = _minutos(ini)
    b = _minutos(fin)
    if b <= a:
        b += 24 * 60
    return a, b


def _se_traslapan(ini1, fin1, ini2, fin2):
    a1, b1 = _rango_minutos(ini1, fin1)
    a2, b2 = _rango_minutos(ini2, fin2)
    return a1 < b2 and a2 < b1


def _calcular_muerto(paros, equipo_nombre, hora_inicio, hora_fin):
    total = 0
    en = equipo_nombre.lower() if equipo_nombre else None
    for p in paros:
        if p.estatus != 'verde':
            continue
        if en and (p.equipo_es or '').lower() != en and (p.equipo_en or '').lower() != en:
            continue
        if hora_fin > hora_inicio:
            if not (p.hora >= hora_inicio and p.hora < hora_fin):
                continue
        else:
            if not (p.hora >= hora_inicio or p.hora < hora_fin):
                continue
        total += p.tiempo_minutos or 0
    return total


def _calcular_kpis_mantenimiento(paros, equipo_nombre, hora_inicio, hora_fin, tiempo_planeado):

    en = equipo_nombre.lower() if equipo_nombre else None
    n_paros = 0
    t_muerto = 0
    for p in paros:
        if p.estatus != 'verde':
            continue
        if p.responsable not in RESPONSABLES_MANT:
            continue
        if en and (p.equipo_es or '').lower() != en and (p.equipo_en or '').lower() != en:
            continue
        if hora_fin > hora_inicio:
            if not (p.hora >= hora_inicio and p.hora < hora_fin):
                continue
        else:
            if not (p.hora >= hora_inicio or p.hora < hora_fin):
                continue
        n_paros += 1
        t_muerto += p.tiempo_minutos or 0

    if n_paros == 0:
        return {'mttr': None, 'mtbf': round(tiempo_planeado / 60, 1), 'disponibilidad': None, 'n_paros': 0, 't_muerto': 0}

    mttr = round(t_muerto / n_paros, 1)
    mtbf = round((tiempo_planeado - t_muerto) / n_paros / 60, 1)
    disp = round(100 - (t_muerto / tiempo_planeado * 100), 1) if tiempo_planeado else None

    return {'mttr': mttr, 'mtbf': mtbf, 'disponibilidad': disp, 'n_paros': n_paros, 't_muerto': t_muerto}


def _construir_registros_data(regs, paros_area):
    """A partir de RegistroProduccion ya filtrados (de una misma área+fecha) y
    los Paro de esa área+fecha, arma las filas con sus KPIs calculados
    (downtime, MTTR, MTBF, disponibilidad...). Usado por la vista de
    Registro de producción y por su exportación a Excel."""
    registros_data = []
    for reg in regs:
        equipo_nombre = reg.equipo or ''
        muerto    = _calcular_muerto(paros_area, equipo_nombre, reg.hora_inicio, reg.hora_fin)
        planeado  = reg.tiempo_planeado
        downtime  = round(muerto / planeado * 100, 1) if planeado else 0

        kpis_mant = _calcular_kpis_mantenimiento(
            paros_area,
            equipo_nombre, reg.hora_inicio, reg.hora_fin, planeado
        )

        registros_data.append({
            'id':              reg.id,
            'equipo':          equipo_nombre or 'Área completa',
            'turno':           reg.turno,
            'hora_inicio':     reg.hora_inicio.strftime('%H:%M'),
            'hora_fin':        reg.hora_fin.strftime('%H:%M'),
            'hora_inicio_raw': reg.hora_inicio.strftime('%H:%M'),
            'hora_fin_raw':    reg.hora_fin.strftime('%H:%M'),
            'numero_molde':    reg.numero_molde,
            'nombre_molde':    reg.nombre_molde,
            'planeado':        planeado,
            'muerto':          muerto,
            'downtime':        downtime,
            't_muerto_mant':   kpis_mant['t_muerto'],
            'mttr':            kpis_mant['mttr'],
            'mtbf':            kpis_mant['mtbf'],
            'disponibilidad':  round(100 - downtime, 1) if planeado else None,
            'n_paros_mant':    kpis_mant['n_paros'],
        })
    return registros_data


@login_required
def registro_produccion(request):
    perfil   = get_perfil(request.user)
    es_admin = request.user.is_superuser or (perfil and perfil.es_admin)

    if es_admin:
        areas = list(Area.objects.all())
    else:
        areas = list(perfil.areas_produccion.all()) if perfil else []

    fecha_str = request.GET.get('fecha', date.today().strftime('%Y-%m-%d'))
    turno     = request.GET.get('turno', '')

    try:
        fecha = date.fromisoformat(fecha_str)
    except ValueError:
        fecha = date.today()

    registros_qs = RegistroProduccion.objects.filter(fecha=fecha, area__in=areas)
    if turno in ('1', '2'):
        registros_qs = registros_qs.filter(turno=int(turno))
    registros_list = list(registros_qs)

    equipos_por_area = defaultdict(list)
    for ce in CatalogoEquipo.objects.filter(area__in=areas).values('area_id', 'codigo', 'equipo'):
        equipos_por_area[ce['area_id']].append({'codigo': ce['codigo'], 'equipo': ce['equipo']})

    area_moldes_id = _area_moldes_id()
    moldes_area = []
    if area_moldes_id and any(a.id == area_moldes_id for a in areas):
        moldes_area = list(
            CatalogoMolde.objects.filter(area_id=area_moldes_id)
            .order_by('numero_molde')
            .values('numero_molde', 'nombre_molde')
        )

    paros_por_area = defaultdict(list)
    for p in Paro.objects.filter(area__in=areas, fecha=fecha):
        paros_por_area[p.area_id].append(p)

    datos_areas = []
    for area in areas:
        regs             = [r for r in registros_list if r.area_id == area.id]
        equipos_catalogo = equipos_por_area[area.id]
        paros_area       = paros_por_area[area.id]

        registros_data = _construir_registros_data(regs, paros_area)
        total_planeado = sum(r['planeado'] for r in registros_data)
        total_muerto   = sum(r['muerto']   for r in registros_data)

        equipo_unicos      = len(set(r['equipo'] for r in registros_data))
        t_muerto_mant_area = sum(r['t_muerto_mant'] for r in registros_data)
        n_paros_mant_area  = sum(r['n_paros_mant']  for r in registros_data)
        downtime_area      = round(total_muerto / total_planeado * 100, 2) if total_planeado else None
        mttr_area          = round(t_muerto_mant_area / n_paros_mant_area, 1) if n_paros_mant_area else None
        mtbf_area          = round(((total_planeado - t_muerto_mant_area) / n_paros_mant_area / 60) / equipo_unicos, 2) if n_paros_mant_area else round(total_planeado / 60 / max(equipo_unicos, 1), 2)
        disp_area          = round(100 - downtime_area, 1) if downtime_area is not None else None

        es_area_moldes = area.id == area_moldes_id

        datos_areas.append({
            'area':               area,
            'registros':          registros_data,
            'equipos':            equipos_catalogo,
            'tiene_equipos':      len(equipos_catalogo) > 0,
            'es_area_moldes':     es_area_moldes,
            'moldes':             moldes_area if es_area_moldes else [],
            'n_columnas':         15 if es_area_moldes else 13,
            'n_registros':        len(registros_data),
            'total_planeado':     total_planeado,
            'total_muerto':       total_muerto,
            'downtime_area':      downtime_area,
            'disp_area':          disp_area,
            'mttr_area':          mttr_area,
            'mtbf_area':          mtbf_area,
            't_muerto_mant_area': t_muerto_mant_area,
            'n_paros_mant_area':  n_paros_mant_area,
        })

    equipos_por_area_json = {str(area_id): lista for area_id, lista in equipos_por_area.items()}

    return render(request, 'paros_app/registro_produccion.html', {
        'datos_areas':           datos_areas,
        'fecha':                 fecha.strftime('%Y-%m-%d'),
        'turno':                 turno,
        'area_moldes_id':        area_moldes_id,
        'moldes_area':           moldes_area,
        'equipos_por_area_json': equipos_por_area_json,
    })


@login_required
@require_POST
def agregar_registro(request):
    try:
        data         = json.loads(request.body)
        area_id      = data.get('area_id')
        equipo       = data.get('equipo', '').strip()
        fecha_str    = data.get('fecha')
        turno        = int(data.get('turno', 1))
        hora_ini     = data.get('hora_inicio')
        hora_fin     = data.get('hora_fin')
        numero_molde = data.get('numero_molde', '').strip()
        nombre_molde = data.get('nombre_molde', '').strip()

        area  = Area.objects.get(id=area_id)
        fecha = date.fromisoformat(fecha_str)

        from datetime import time as _time
        hi = _time.fromisoformat(hora_ini)
        hf = _time.fromisoformat(hora_fin)

        existentes = RegistroProduccion.objects.filter(area=area, equipo=equipo, fecha=fecha, turno=turno)
        if area.id == _area_moldes_id():
            for existente in existentes:
                if _se_traslapan(hi, hf, existente.hora_inicio, existente.hora_fin):
                    return JsonResponse({'ok': False, 'error': f'El horario se traslapa con un registro existente ({existente.hora_inicio.strftime("%H:%M")}–{existente.hora_fin.strftime("%H:%M")}).'}, status=400)
        else:
            if existentes.exists():
                return JsonResponse({'ok': False, 'error': 'Ya existe un registro para este equipo y turno en esta fecha.'}, status=400)

        equipo_es, equipo_en = _normalizar_equipo(area, equipo)
        reg = RegistroProduccion.objects.create(
            area=area, equipo_es=equipo_es, equipo_en=equipo_en, fecha=fecha, turno=turno,
            hora_inicio=hi, hora_fin=hf, numero_molde=numero_molde, nombre_molde=nombre_molde,
        )

        paros_qs = list(Paro.objects.filter(area=area, fecha=fecha, turno=turno))
        muerto   = _calcular_muerto(paros_qs, equipo, hi, hf)
        planeado = reg.tiempo_planeado
        downtime = round(muerto / planeado * 100, 1) if planeado else 0

        return JsonResponse({'ok': True, 'id': reg.id, 'planeado': planeado, 'muerto': muerto, 'downtime': downtime})
    except Exception as e:
        error = str(e)
        if 'unique' in error.lower() or 'duplicate' in error.lower():
            error = 'Ya existe un registro para este equipo y turno en esta fecha.'
        return JsonResponse({'ok': False, 'error': error}, status=400)


@login_required
@require_POST
def eliminar_registro(request, registro_id):
    try:
        RegistroProduccion.objects.filter(id=registro_id).delete()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
@require_POST
def actualizar_registro(request, registro_id):
    try:
        data         = json.loads(request.body)
        hora_ini     = data.get('hora_inicio')
        hora_fin     = data.get('hora_fin')
        turno        = data.get('turno')
        equipo       = data.get('equipo')
        numero_molde = data.get('numero_molde')
        nombre_molde = data.get('nombre_molde')


        from datetime import time as _time
        reg = RegistroProduccion.objects.get(id=registro_id)

        equipo_final = equipo if equipo is not None else reg.equipo
        turno_final  = int(turno) if turno else reg.turno
        hi_final     = _time.fromisoformat(hora_ini) if hora_ini else reg.hora_inicio
        hf_final     = _time.fromisoformat(hora_fin) if hora_fin else reg.hora_fin

        existentes = RegistroProduccion.objects.filter(
            area=reg.area, equipo=equipo_final, fecha=reg.fecha, turno=turno_final
        ).exclude(id=reg.id)
        if reg.area_id == _area_moldes_id():
            for existente in existentes:
                if _se_traslapan(hi_final, hf_final, existente.hora_inicio, existente.hora_fin):
                    return JsonResponse({'ok': False, 'error': f'El horario se traslapa con un registro existente ({existente.hora_inicio.strftime("%H:%M")}–{existente.hora_fin.strftime("%H:%M")}).'}, status=400)
        else:
            if existentes.exists():
                return JsonResponse({'ok': False, 'error': 'Ya existe un registro para ese equipo y turno en esta fecha.'}, status=400)

        if equipo is not None:
            equipo_es, equipo_en = _normalizar_equipo(reg.area, equipo_final)
            reg.equipo_es = equipo_es
            reg.equipo_en = equipo_en

        reg.turno       = turno_final
        reg.hora_inicio = hi_final
        reg.hora_fin    = hf_final
        if numero_molde is not None: reg.numero_molde = numero_molde
        if nombre_molde is not None: reg.nombre_molde = nombre_molde

        reg.save()

        paros_qs = list(Paro.objects.filter(area=reg.area, fecha=reg.fecha, turno=reg.turno))
        muerto   = _calcular_muerto(paros_qs, reg.equipo, reg.hora_inicio, reg.hora_fin)
        planeado = reg.tiempo_planeado
        downtime = round(muerto / planeado * 100, 1) if planeado else 0

        return JsonResponse({'ok': True, 'planeado': planeado, 'muerto': muerto, 'downtime': downtime})
    except Exception as e:
        error = str(e)
        if 'unique' in error.lower() or 'UniqueViolation' in error or 'duplicada' in error:
            error = 'Ya existe un registro para ese equipo y turno en esta fecha.'
        return JsonResponse({'ok': False, 'error': error}, status=400)


@login_required
@require_POST
def actualizar_orden(request):
    try:
        data  = json.loads(request.body)
        orden = data.get('orden', [])
        posicion = {int(reg_id): i for i, reg_id in enumerate(orden)}
        registros = list(RegistroProduccion.objects.filter(id__in=posicion.keys()))
        for reg in registros:
            reg.orden = posicion[reg.id]
        RegistroProduccion.objects.bulk_update(registros, ['orden'])
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


def _resolver_filtros_export(request):
    """Resuelve permiso + áreas + rango de fechas + turno para la exportación
    de registros de producción. Devuelve (areas, fecha_desde, fecha_hasta, turno)
    o None si el usuario no tiene permiso. Usado por exportar_registros_produccion
    y contar_registros_produccion."""
    perfil   = get_perfil(request.user)
    es_admin = request.user.is_superuser or (perfil and perfil.es_admin)

    if es_admin:
        areas = list(Area.objects.all())
    else:
        areas = list(perfil.areas_produccion.all()) if perfil else []

    if not areas:
        return None

    hoy = date.today()
    try:
        fecha_desde = date.fromisoformat(request.GET.get('fecha_desde', ''))
    except ValueError:
        fecha_desde = hoy
    try:
        fecha_hasta = date.fromisoformat(request.GET.get('fecha_hasta', ''))
    except ValueError:
        fecha_hasta = hoy
    if fecha_hasta < fecha_desde:
        fecha_desde, fecha_hasta = fecha_hasta, fecha_desde

    area_id = request.GET.get('area_id', '')
    if area_id:
        areas = [a for a in areas if str(a.id) == area_id]
        if not areas:
            return None

    turno = request.GET.get('turno', '')

    return areas, fecha_desde, fecha_hasta, turno


@login_required
def contar_registros_produccion(request):
    resultado = _resolver_filtros_export(request)
    if resultado is None:
        return HttpResponseForbidden('Sin permiso para exportar.')
    areas, fecha_desde, fecha_hasta, turno = resultado

    qs = RegistroProduccion.objects.filter(
        area__in=areas, fecha__gte=fecha_desde, fecha__lte=fecha_hasta
    )
    if turno in ('1', '2'):
        qs = qs.filter(turno=int(turno))
    return JsonResponse({'count': qs.count()})


@login_required
def exportar_registros_produccion(request):
    resultado = _resolver_filtros_export(request)
    if resultado is None:
        return HttpResponseForbidden('Sin permiso para exportar.')
    areas, fecha_desde, fecha_hasta, turno = resultado

    registros_qs = RegistroProduccion.objects.filter(
        area__in=areas, fecha__gte=fecha_desde, fecha__lte=fecha_hasta
    ).order_by('fecha', 'area_id', 'orden')
    if turno in ('1', '2'):
        registros_qs = registros_qs.filter(turno=int(turno))
    registros_list = list(registros_qs)

    paros_qs = Paro.objects.filter(area__in=areas, fecha__gte=fecha_desde, fecha__lte=fecha_hasta)

    paros_por_area_fecha = defaultdict(list)
    for p in paros_qs:
        paros_por_area_fecha[(p.area_id, p.fecha)].append(p)

    regs_por_area_fecha = defaultdict(list)
    for r in registros_list:
        regs_por_area_fecha[(r.area_id, r.fecha)].append(r)

    areas_por_id = {a.id: a for a in areas}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registros de producción'
    cabeceras = [
        'Área', 'Fecha', 'Equipo', 'Turno', 'Inicio', 'Fin',
        'No. molde', 'Nombre de molde', 'Planeado (min)', 'T. muerto (min)',
        'Downtime %', 'Disponibilidad %', 'T. muerto mant. (min)', 'MTTR (min)', 'MTBF (h)', 'No. Paros mant.',
    ]
    ws.append(cabeceras)

    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_font = Font(bold=True, color='FFFFFF')
    center = Alignment(horizontal='center', vertical='center')
    for col in range(1, len(cabeceras) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for (area_id_k, fecha_k), regs in sorted(regs_por_area_fecha.items(), key=lambda kv: (kv[0][1], kv[0][0])):
        area = areas_por_id.get(area_id_k)
        if not area:
            continue
        paros_area     = paros_por_area_fecha[(area_id_k, fecha_k)]
        registros_data = _construir_registros_data(regs, paros_area)
        for reg, rd in zip(regs, registros_data):
            ws.append([
                area.nombre_es or area.nombre,
                fecha_k.strftime('%d/%m/%Y'),
                rd['equipo'],
                reg.get_turno_display(),
                reg.hora_inicio.strftime('%H:%M'),
                reg.hora_fin.strftime('%H:%M'),
                reg.numero_molde or '—',
                reg.nombre_molde or '—',
                rd['planeado'],
                rd['muerto'],
                rd['downtime'],
                rd['disponibilidad'] if rd['disponibilidad'] is not None else '',
                rd['t_muerto_mant'],
                rd['mttr'] if rd['mttr'] is not None else '',
                rd['mtbf'] if rd['mtbf'] is not None else '',
                rd['n_paros_mant'],
            ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = center
            if cell.column == 2:
                cell.number_format = '@'

    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ''
                if len(val) > max_len:
                    max_len = len(val)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    if request.GET.get('area_id', '') and areas:
        filename = f"Registros_Produccion_{areas[0].nombre.replace(' ', '_')}.xlsx"
    else:
        filename = f"Registros_Produccion_{fecha_desde.strftime('%Y%m%d')}_a_{fecha_hasta.strftime('%Y%m%d')}.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
