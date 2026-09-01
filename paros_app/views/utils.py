from datetime import datetime, date as _date


def _aplicar_filtros(qs, get_params):
    area_id       = get_params.get('area')
    fecha_desde   = get_params.get('fecha_desde')
    fecha_hasta   = get_params.get('fecha_hasta')
    turno         = get_params.get('turno')
    busqueda      = get_params.get('q', '').strip()
    filtro_resp   = get_params.get('resp', '').strip()
    filtro_equipo = get_params.get('equipo', '').strip()

    if area_id:
        qs = qs.filter(area_id=area_id)
    semana_activa = get_params.get('semana', '').strip()
    if not semana_activa:
        if fecha_desde:
            try:
                qs = qs.filter(fecha__gte=datetime.strptime(fecha_desde, '%Y-%m-%d').date())
            except ValueError:
                pass
        if fecha_hasta:
            try:
                qs = qs.filter(fecha__lte=datetime.strptime(fecha_hasta, '%Y-%m-%d').date())
            except ValueError:
                pass
    if turno in ('1', '2'):
        qs = qs.filter(turno=int(turno))
    if busqueda:
        qs = qs.filter(falla__icontains=busqueda)
    if filtro_resp:
        qs = qs.filter(responsable__icontains=filtro_resp)
    if filtro_equipo:
        qs = qs.filter(equipo__icontains=filtro_equipo)

    semana = get_params.get('semana', '').strip()
    if semana:
        try:
            semana_int = int(semana)
            año_str = get_params.get('fecha_desde', '')
            try:
                año_int = int(año_str[:4]) if año_str else _date.today().year
            except Exception:
                año_int = _date.today().year
            lunes   = _date.fromisocalendar(año_int, semana_int, 1)
            domingo = _date.fromisocalendar(año_int, semana_int, 7)
            qs = qs.filter(fecha__gte=lunes, fecha__lte=domingo)
        except (ValueError, TypeError):
            pass
    return qs


ORDEN_PERMITIDO = {
    'fecha': 'fecha', '-fecha': '-fecha',
    'hora': 'hora', '-hora': '-hora',
    'tiempo_minutos': 'tiempo_minutos', '-tiempo_minutos': '-tiempo_minutos',
}


def _aplicar_orden(qs, get_params, default=None):
    """Ordena por el campo pedido en ?orden=; si no es válido, usa `default`."""
    campo = ORDEN_PERMITIDO.get(get_params.get('orden', ''))
    if campo:
        return qs.order_by(campo)
    if default:
        return qs.order_by(*default)
    return qs


def _parse_fecha(s):
    """
    Parsea fecha desde string dd/mm/yyyy, dd/mm/yy, dd-mm-yy
    o desde objeto datetime/date de openpyxl (cuando Excel convierte la celda).
    """
    from datetime import date as _d, datetime as _dt
    if isinstance(s, (_dt, _d)):
        if isinstance(s, _dt):
            return s.date()
        return s
    s_original = str(s).strip()
    s = s_original.replace('-', '/').replace('.', '/')
    partes = s.split('/')
    if len(partes) != 3:
        raise ValueError(f"Fecha '{s_original}' no reconocida (usa el formato dd/mm/aaaa)")
    try:
        dia  = int(partes[0])
        mes  = int(partes[1])
        anio = int(partes[2])
    except ValueError:
        raise ValueError(f"Fecha '{s_original}' no reconocida (usa el formato dd/mm/aaaa)")
    if anio < 100:
        anio += 2000
    if mes > 12:
        raise ValueError(f"Fecha '{s_original}' inválida — el mes '{mes}' no existe (máximo 12). Usa el formato dd/mm/aaaa, no mm/dd/aaaa")
    try:
        return _d(anio, mes, dia)
    except ValueError:
        raise ValueError(f"Fecha '{s_original}' inválida — revisa que sea dd/mm/aaaa (día/mes/año)")


import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse


def _excel_response(filename):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _estilo_cabecera(ws, cabeceras, col_widths):
    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_font = Font(bold=True, color='FFFFFF')
    ws.append(cabeceras)
    for col in range(1, len(cabeceras) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w