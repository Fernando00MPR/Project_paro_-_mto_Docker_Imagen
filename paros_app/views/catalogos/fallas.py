import csv
import io
import openpyxl

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_http_methods

from ...models import Area, CatalogoFalla
from ...views.utils import _excel_response, _estilo_cabecera
from login_app.permisos import permiso_requerido, get_perfil
from openpyxl.styles import Alignment

@login_required
@permiso_requerido('ver_catalogos')
def catalogo_fallas_general(request):
    perfil = get_perfil(request.user)
    areas  = Area.objects.prefetch_related('catalogo_fallas').all()
    if not (request.user.is_superuser or (perfil and perfil.es_admin)):
        if perfil and perfil.areas_permitidas.exists():
            areas = areas.filter(id__in=perfil.areas_permitidas.all())
        else:
            areas = areas.none()
    puede_gestionar = request.user.is_superuser or (perfil and (perfil.es_admin or perfil.gestionar_catalogos))
    return render(request, 'paros_app/fallas/catalogo_fallas_general.html', {
        'areas':           areas,
        'puede_gestionar': puede_gestionar,
    })


@login_required
@permiso_requerido('ver_catalogos')
def catalogo_fallas(request, area_id):
    perfil = get_perfil(request.user)
    if not (request.user.is_superuser or (perfil and perfil.es_admin)):
        if perfil and perfil.areas_permitidas.exists():
            if not perfil.areas_permitidas.filter(id=area_id).exists():
                messages.error(request, "No tienes acceso al catálogo de esta área.")
                return redirect('paros:catalogo_fallas_general')
    area = get_object_or_404(Area, id=area_id)
    qs   = CatalogoFalla.objects.filter(area=area).order_by('codigo')
    q    = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(Q(nombre__icontains=q) | Q(codigo__icontains=q))
    paginator = Paginator(qs, 50)
    fallas    = paginator.get_page(request.GET.get('page', 1))
    puede_gestionar = request.user.is_superuser or (perfil and (perfil.es_admin or perfil.gestionar_catalogos))
    return render(request, 'paros_app/fallas/catalogo_fallas.html', {
        'area':            area,
        'fallas':          fallas,
        'page_obj':        fallas,
        'areas':           Area.objects.all(),
        'q':               q,
        'puede_gestionar': puede_gestionar,
    })


@login_required
@permiso_requerido('gestionar_catalogos')
@require_http_methods(["POST"])
def eliminar_falla(request, falla_id):
    falla = get_object_or_404(CatalogoFalla, id=falla_id)
    falla.delete()
    return redirect('paros:catalogo_fallas_general')


@login_required
@permiso_requerido('gestionar_catalogos')
@require_http_methods(["POST"])
def limpiar_fallas_area(request, area_id):
    area = get_object_or_404(Area, id=area_id)
    CatalogoFalla.objects.filter(area=area).delete()
    messages.success(request, f"Catálogo de fallas de '{area.nombre}' eliminado.")
    return redirect('paros:catalogo_fallas_general')


@login_required
@permiso_requerido('gestionar_catalogos')
def importar_fallas_v2(request):
    """Importación con 3 columnas en Excel: código | área | falla."""
    errores = []
    resumen = None

    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        modo    = request.POST.get('modo', 'agregar')
        if not archivo:
            errores.append("Selecciona un archivo Excel.")
        else:
            nombre = archivo.name.lower()
            filas  = []
            try:
                if nombre.endswith(('.xlsx', '.xls')):
                    wb   = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
                    ws   = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    for i, row in enumerate(rows[1:], start=2):
                        if not any(row): continue
                        codigo       = str(row[0] or '').strip() if len(row) > 0 else ''
                        nombre_area  = str(row[1] or '').strip() if len(row) > 1 else ''
                        falla_es     = str(row[2] or '').strip() if len(row) > 2 else ''
                        falla_en     = str(row[3] or '').strip() if len(row) > 3 else ''
                        tipo_falla   = str(row[4] or '').strip() if len(row) > 4 else ''
                        tipo_falla_en = str(row[5] or '').strip() if len(row) > 5 else ''
                        area_origen  = str(row[6] or '').strip() if len(row) > 6 else ''
                        if not codigo or not nombre_area or not falla_es:
                            errores.append(f"Fila {i}: código, área y falla (ES) son obligatorios.")
                            continue
                        filas.append((codigo, nombre_area, falla_es, falla_en, tipo_falla, tipo_falla_en, area_origen))
                elif nombre.endswith('.csv'):
                    contenido = archivo.read().decode('utf-8-sig')
                    reader    = csv.DictReader(io.StringIO(contenido))
                    for i, row in enumerate(reader, start=2):
                        codigo        = (row.get('codigo')       or row.get('código')      or '').strip()
                        nombre_area   = (row.get('area')         or row.get('área')        or '').strip()
                        falla_es      = (row.get('falla_es')     or row.get('falla')       or '').strip()
                        falla_en      = (row.get('falla_en')     or '').strip()
                        tipo_falla    = (row.get('tipo_falla')   or '').strip()
                        tipo_falla_en = (row.get('tipo_falla_en') or '').strip()
                        area_origen   = (row.get('area_origen')  or row.get('area origen') or '').strip()
                        if not codigo or not nombre_area or not falla_es:
                            errores.append(f"Fila {i}: datos incompletos.")
                            continue
                        filas.append((codigo, nombre_area, falla_es, falla_en, tipo_falla, tipo_falla_en, area_origen))
                else:
                    errores.append("Solo se aceptan archivos .xlsx, .xls o .csv")
            except Exception as e:
                errores.append(f"Error al leer el archivo: {e}")

            if filas and not errores:
                creados = actualizados = omitidos = 0
                areas_borradas   = set()
                errores_guardado = []
                try:
                    with transaction.atomic():
                        for codigo, nombre_area, falla_es, falla_en, tipo_falla, tipo_falla_en, area_origen in filas:
                            try:
                                area = Area.objects.get(nombre_es__iexact=nombre_area)
                            except Area.DoesNotExist:
                                try:
                                    area = Area.objects.get(nombre__iexact=nombre_area)
                                except Area.DoesNotExist:
                                    errores_guardado.append(f"Área '{nombre_area}' no existe.")
                                    omitidos += 1
                                    continue
                            if modo == 'reemplazar' and area.id not in areas_borradas:
                                CatalogoFalla.objects.filter(area=area).delete()
                                areas_borradas.add(area.id)
                            _, created = CatalogoFalla.objects.update_or_create(
                                area=area, codigo=codigo,
                                defaults={
                                    'nombre': falla_es, 'nombre_es': falla_es, 'nombre_en': falla_en,
                                    'tipo_falla': tipo_falla, 'tipo_falla_es': tipo_falla, 'tipo_falla_en': tipo_falla_en,
                                    'area_origen': area_origen
                                }
                            )
                            if created: creados += 1
                            else: actualizados += 1
                        if errores_guardado:
                            raise Exception("Errores en importación")
                except Exception:
                    errores.extend(errores_guardado)
                    creados = actualizados = 0

                if not errores:
                    resumen = {
                        'creados':      creados,
                        'actualizados': actualizados,
                        'omitidos':     omitidos,
                        'total':        creados + actualizados,
                    }

    areas = Area.objects.all()
    return render(request, 'paros_app/fallas/importar_fallas.html', {
        'areas': areas, 'errores': errores, 'resumen': resumen,
    })


@login_required
@permiso_requerido('gestionar_catalogos')
def importar_fallas_por_area(request, area_id):
    area    = get_object_or_404(Area, id=area_id)
    errores = []
    resumen = None

    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        modo    = request.POST.get('modo', 'agregar')
        if not archivo:
            errores.append("Selecciona un archivo.")
        else:
            filas = []
            try:
                nombre = archivo.name.lower()
                if nombre.endswith(('.xlsx', '.xls')):
                    wb   = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
                    ws   = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        errores.append("El archivo está vacío.")
                    else:
                        headers = [str(c).strip().lower() if c else '' for c in rows[0]]
                        def col(names):
                            for n in names:
                                if n in headers: return headers.index(n)
                            return None
                        idx_cod      = col(['código', 'codigo'])
                        idx_fall_es  = col(['falla_es', 'falla (español)', 'falla (es)', 'falla'])
                        idx_fall_en  = col(['falla_en', 'falla (english)', 'falla (en)', 'fault (english)'])
                        idx_tipo_es  = col(['tipo_falla_es', 'tipo falla', 'tipo de falla', 'tipo de falla (es)', 'tipo falla (es)'])
                        idx_tipo_en  = col(['tipo_falla_en', 'tipo falla (en)', 'tipo de falla (en)', 'fault type', 'tipo falla en'])
                        idx_orig     = col(['área de origen', 'area de origen', 'area_origen', 'area origen'])
                        idx_area     = col(['área', 'area'])
                        if idx_cod is None or idx_fall_es is None:
                            errores.append("El archivo debe tener columnas 'Código' y 'Falla (ES)'.")
                        else:
                            for i, row in enumerate(rows[1:], start=2):
                                if not any(row): continue
                                if idx_area is not None and idx_area < len(row):
                                    area_col = str(row[idx_area] or '').strip()
                                    if area_col and area_col.lower() != area.nombre.lower():
                                        continue
                                codigo      = str(row[idx_cod]     or '').strip()
                                falla_es    = str(row[idx_fall_es] or '').strip()
                                falla_en    = str(row[idx_fall_en] or '').strip() if idx_fall_en  is not None and idx_fall_en  < len(row) else ''
                                tipo_falla  = str(row[idx_tipo_es] or '').strip() if idx_tipo_es  is not None and idx_tipo_es  < len(row) else ''
                                tipo_falla_en = str(row[idx_tipo_en] or '').strip() if idx_tipo_en is not None and idx_tipo_en < len(row) else ''
                                area_origen = str(row[idx_orig]    or '').strip() if idx_orig     is not None and idx_orig     < len(row) else ''
                                if not codigo or not falla_es:
                                    errores.append(f"Fila {i}: código y falla (ES) son obligatorios.")
                                    continue
                                filas.append((codigo, falla_es, falla_en, tipo_falla, tipo_falla_en, area_origen))
                elif nombre.endswith('.csv'):
                    contenido = archivo.read().decode('utf-8-sig')
                    reader = csv.DictReader(io.StringIO(contenido))
                    for i, row in enumerate(reader, start=2):
                        row_lower = {k.strip().lower(): v for k, v in row.items()}
                        area_col = (row_lower.get('área') or row_lower.get('area') or '').strip()
                        if area_col and area_col.lower() != area.nombre.lower():
                            continue
                        falla_es      = (row_lower.get('falla_es')    or row_lower.get('falla') or '').strip()
                        falla_en      = (row_lower.get('falla_en')    or '').strip()
                        tipo_falla    = (row_lower.get('tipo_falla_es') or row_lower.get('tipo_falla') or '').strip()
                        tipo_falla_en = (row_lower.get('tipo_falla_en') or '').strip()
                        area_origen   = (row_lower.get('área de origen') or row_lower.get('area de origen') or row_lower.get('area_origen') or '').strip()
                        if not codigo or not falla_es:
                            errores.append(f"Fila {i}: datos incompletos.")
                            continue
                        filas.append((codigo, falla_es, falla_en, tipo_falla, tipo_falla_en, area_origen))
                else:
                    errores.append("Solo .xlsx, .xls o .csv")
            except Exception as e:
                errores.append(f"Error al leer el archivo: {e}")

            if filas and not errores:
                if modo == 'reemplazar':
                    CatalogoFalla.objects.filter(area=area).delete()
                creados = actualizados = 0
                for codigo, falla_es, falla_en, tipo_falla, tipo_falla_en, area_origen in filas:
                    _, created = CatalogoFalla.objects.update_or_create(
                        area=area, codigo=codigo,
                        defaults={
                            'nombre': falla_es, 'nombre_es': falla_es, 'nombre_en': falla_en,
                            'tipo_falla': tipo_falla, 'tipo_falla_es': tipo_falla, 'tipo_falla_en': tipo_falla_en,
                            'area_origen': area_origen
                        }
                    )
                    if created: creados += 1
                    else: actualizados += 1
                resumen = {'creados': creados, 'actualizados': actualizados, 'area': area.nombre}

    return render(request, 'paros_app/fallas/importar_fallas_por_area.html', {
        'area': area, 'errores': errores, 'resumen': resumen,
    })


@login_required
def descargar_plantilla_fallas_v2(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="Plantilla_Fallas_General.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow(['codigo', 'area', 'falla_es', 'falla_en', 'tipo_falla', 'tipo_falla_en', 'area_origen'])
    writer.writerow(['F-001', 'Area A', 'Falla A', 'Fault A', 'Tipo A', 'Type A', 'Area Origen A'])
    writer.writerow(['F-002', 'Area B', 'Falla B', 'Fault B', 'Tipo B', 'Type B', 'Area Origen B'])
    writer.writerow(['F-003', 'Area C', 'Falla C', 'Fault C', 'Tipo C', 'Type C', 'Area Origen C'])
    return response


@login_required
@permiso_requerido('ver_catalogos')
def exportar_fallas(request, area_id=None):
    qs = CatalogoFalla.objects.select_related('area').all()
    if area_id:
        area  = get_object_or_404(Area, id=area_id)
        qs    = qs.filter(area=area)
        fname = f'Fallas_{area.nombre.replace(" ", "_")}.xlsx'
    else:
        fname = 'Fallas_General.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fallas'
    if area_id:
        _estilo_cabecera(ws, ['Código', 'Falla (ES)', 'Falla (EN)', 'Tipo de falla (ES)', 'Tipo de falla (EN)', 'Área de origen'], [14, 38, 38, 22, 22, 20])
        for f in qs.order_by('area__nombre', 'codigo'):
            ws.append([f.codigo, f.nombre_es or f.nombre, f.nombre_en or '', f.tipo_falla_es or f.tipo_falla or '', f.tipo_falla_en or '', f.area_origen or ''])
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    else:
        _estilo_cabecera(ws, ['Código', 'Área', 'Falla (ES)', 'Falla (EN)', 'Tipo de falla (ES)', 'Tipo de falla (EN)', 'Área de origen'], [14, 22, 38, 38, 22, 22, 20])
        for f in qs.order_by('area__nombre', 'codigo'):
            ws.append([f.codigo, f.area.nombre_es or f.area.nombre, f.nombre_es or f.nombre, f.nombre_en or '', f.tipo_falla_es or f.tipo_falla or '', f.tipo_falla_en or '', f.area_origen or ''])
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    response = _excel_response(fname)
    wb.save(response)
    return response


@login_required
@permiso_requerido('gestionar_catalogos')
def agregar_falla(request, area_id):
    area = get_object_or_404(Area, id=area_id)
    if request.method == 'POST':
        codigo        = request.POST.get('codigo', '').strip()
        nombre_es     = request.POST.get('nombre_es', '').strip()
        nombre_en     = request.POST.get('nombre_en', '').strip()
        tipo_falla_es = request.POST.get('tipo_falla_es', '').strip()
        tipo_falla_en = request.POST.get('tipo_falla_en', '').strip()
        area_origen   = request.POST.get('area_origen', '').strip()

        if not codigo or not nombre_es:
            messages.error(request, "Código y nombre en español son obligatorios.")
        elif CatalogoFalla.objects.filter(area=area, codigo=codigo).exists():
            messages.error(request, f"Ya existe una falla con el código '{codigo}' en esta área.")
        else:
            CatalogoFalla.objects.create(
                area=area,
                codigo=codigo,
                nombre=nombre_es,
                nombre_es=nombre_es,
                nombre_en=nombre_en,
                tipo_falla=tipo_falla_es,
                tipo_falla_es=tipo_falla_es,
                tipo_falla_en=tipo_falla_en,
                area_origen=area_origen,
            )
            messages.success(request, f"Falla '{nombre_es}' agregada correctamente.")

    return redirect('paros:catalogo_fallas_general')

@login_required
@permiso_requerido('gestionar_catalogos')
def editar_falla(request, falla_id):
    falla = get_object_or_404(CatalogoFalla, id=falla_id)
    if request.method == 'POST':
        codigo        = request.POST.get('codigo', '').strip()
        nombre_es     = request.POST.get('nombre_es', '').strip()
        nombre_en     = request.POST.get('nombre_en', '').strip()
        tipo_falla_es = request.POST.get('tipo_falla_es', '').strip()
        tipo_falla_en = request.POST.get('tipo_falla_en', '').strip()
        area_origen   = request.POST.get('area_origen', '').strip()
        if not codigo or not nombre_es:
            messages.error(request, "Código y nombre en español son obligatorios.")
        elif CatalogoFalla.objects.filter(area=falla.area, codigo=codigo).exclude(id=falla.id).exists():
            messages.error(request, f"Ya existe una falla con el código '{codigo}'.")
        else:
            falla.codigo        = codigo
            falla.nombre        = nombre_es
            falla.nombre_es     = nombre_es
            falla.nombre_en     = nombre_en
            falla.tipo_falla    = tipo_falla_es
            falla.tipo_falla_es = tipo_falla_es
            falla.tipo_falla_en = tipo_falla_en
            falla.area_origen   = area_origen
            falla.save()
            messages.success(request, f"Falla '{nombre_es}' actualizada correctamente.")
    return redirect('paros:catalogo_fallas_general')