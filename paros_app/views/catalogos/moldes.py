import csv
import io
import openpyxl

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_http_methods

from ...models import Area, CatalogoMolde, ConfiguracionMoldes
from ...views.utils import _excel_response, _estilo_cabecera
from login_app.permisos import permiso_requerido, get_perfil
from openpyxl.styles import Alignment


def _area_moldes():
    config = ConfiguracionMoldes.objects.select_related('area').first()
    return config.area if config else None


@login_required
@permiso_requerido('ver_catalogos')
def catalogo_moldes_general(request):
    perfil   = get_perfil(request.user)
    es_admin = request.user.is_superuser or (perfil and perfil.es_admin)
    area     = _area_moldes()

    areas = Area.objects.none()
    if area:
        permitido = es_admin or not (perfil and perfil.areas_permitidas.exists()) or perfil.areas_permitidas.filter(id=area.id).exists()
        if permitido:
            areas = Area.objects.filter(id=area.id).prefetch_related('catalogo_moldes').annotate(num_moldes=Count('catalogo_moldes'))

    return render(request, 'paros_app/moldes/catalogo_moldes.html', {
        'areas':           areas,
        'area_configurada': area,
        'puede_gestionar': es_admin or (perfil and perfil.gestionar_catalogos),
        'puede_agregar':   es_admin or (perfil and perfil.agregar_catalogo_molde),
        'puede_editar':    es_admin or (perfil and perfil.editar_catalogo_molde),
        'puede_eliminar':  es_admin or (perfil and perfil.eliminar_catalogo_molde),
    })


@login_required
@permiso_requerido('gestionar_catalogos')
@require_http_methods(["POST"])
def limpiar_moldes_area(request, area_id):
    area = get_object_or_404(Area, id=area_id)
    CatalogoMolde.objects.filter(area=area).delete()
    messages.success(request, f"Catálogo de moldes de '{area.nombre}' eliminado.")
    return redirect('paros:catalogo_moldes')


@login_required
@permiso_requerido('gestionar_catalogos')
def importar_moldes(request):
    area = _area_moldes()
    errores = []
    resumen = None

    if area is None:
        errores.append("No se ha configurado el área de moldes. Contacta a un administrador en el panel de administración.")
        return render(request, 'paros_app/moldes/importar_moldes.html', {
            'area': None, 'errores': errores, 'resumen': resumen,
        })

    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        modo    = request.POST.get('modo', 'agregar')
        if not archivo:
            errores.append("Selecciona un archivo.")
        else:
            filas = []
            try:
                nombre = archivo.name.lower()
                if nombre.endswith(('.xlsx', '.xls')):
                    wb   = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
                    ws   = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    for i, row in enumerate(rows[1:], start=2):
                        if not any(row): continue
                        numero       = str(row[0] or '').strip() if len(row) > 0 else ''
                        nombre_molde = str(row[1] or '').strip() if len(row) > 1 else ''
                        if not numero or not nombre_molde:
                            errores.append(f"Fila {i}: número y nombre de molde son obligatorios.")
                            continue
                        filas.append((numero, nombre_molde))
                elif nombre.endswith('.csv'):
                    contenido = archivo.read().decode('utf-8-sig')
                    reader = csv.DictReader(io.StringIO(contenido))
                    for i, row in enumerate(reader, start=2):
                        row_lower = {k.strip().lower(): v for k, v in row.items()}
                        numero       = (row_lower.get('numero_molde') or row_lower.get('número de molde') or '').strip()
                        nombre_molde = (row_lower.get('nombre_molde') or row_lower.get('nombre de molde')  or '').strip()
                        if not numero or not nombre_molde:
                            errores.append(f"Fila {i}: datos incompletos.")
                            continue
                        filas.append((numero, nombre_molde))
                else:
                    errores.append("Solo se aceptan .xlsx, .xls o .csv")
            except Exception as e:
                errores.append(f"Error al leer el archivo: {e}")

            if filas and not errores:
                if modo == 'reemplazar':
                    CatalogoMolde.objects.filter(area=area).delete()
                creados = actualizados = 0
                for numero, nombre_molde in filas:
                    _, created = CatalogoMolde.objects.update_or_create(
                        area=area, numero_molde=numero,
                        defaults={'nombre_molde': nombre_molde}
                    )
                    if created: creados += 1
                    else: actualizados += 1
                resumen = {'creados': creados, 'actualizados': actualizados, 'area': area.nombre}

    return render(request, 'paros_app/moldes/importar_moldes.html', {
        'area': area, 'errores': errores, 'resumen': resumen,
    })


@login_required
def descargar_plantilla_moldes(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="Plantilla_Moldes.csv"'
    response.write('﻿')
    writer = csv.writer(response)
    writer.writerow(['numero_molde', 'nombre_molde'])
    writer.writerow(['M-001', 'Molde A'])
    writer.writerow(['M-002', 'Molde B'])
    return response


@login_required
@permiso_requerido('eliminar_catalogo_molde')
@require_http_methods(["POST"])
def eliminar_molde(request, molde_id):
    molde = get_object_or_404(CatalogoMolde, id=molde_id)
    molde.delete()
    return redirect('paros:catalogo_moldes')


@login_required
@permiso_requerido('ver_catalogos')
def exportar_moldes(request, area_id=None):
    area = get_object_or_404(Area, id=area_id) if area_id else _area_moldes()
    qs   = CatalogoMolde.objects.filter(area=area).order_by('numero_molde') if area else CatalogoMolde.objects.none()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Moldes'
    fname = f'Moldes_{area.nombre_es or area.nombre}.xlsx'.replace(" ", "_") if area else 'Moldes.xlsx'
    _estilo_cabecera(ws, ['Número de molde', 'Nombre de molde'], [16, 40])
    for m in qs:
        ws.append([m.numero_molde, m.nombre_molde])
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    response = _excel_response(fname)
    wb.save(response)
    return response


@login_required
@permiso_requerido('agregar_catalogo_molde')
def agregar_moldes(request, area_id):
    area = get_object_or_404(Area, id=area_id)
    if request.method == 'POST':
        numero       = request.POST.get('numero_molde', '').strip()
        nombre_molde = request.POST.get('nombre_molde', '').strip()
        if not numero or not nombre_molde:
            messages.error(request, "Número y nombre de molde son obligatorios.")
        elif CatalogoMolde.objects.filter(area=area, numero_molde=numero).exists():
            messages.error(request, f"Ya existe un molde con el número '{numero}' en esta área.")
        else:
            CatalogoMolde.objects.create(area=area, numero_molde=numero, nombre_molde=nombre_molde)
            messages.success(request, f"Molde '{nombre_molde}' agregado correctamente.")
    return redirect('paros:catalogo_moldes')


@login_required
@permiso_requerido('editar_catalogo_molde')
def editar_molde(request, molde_id):
    molde = get_object_or_404(CatalogoMolde, id=molde_id)
    if request.method == 'POST':
        numero       = request.POST.get('numero_molde', '').strip()
        nombre_molde = request.POST.get('nombre_molde', '').strip()
        if not numero or not nombre_molde:
            messages.error(request, "Número y nombre de molde son obligatorios.")
        elif CatalogoMolde.objects.filter(area=molde.area, numero_molde=numero).exclude(id=molde.id).exists():
            messages.error(request, f"Ya existe un molde con el número '{numero}'.")
        else:
            molde.numero_molde = numero
            molde.nombre_molde = nombre_molde
            molde.save()
            messages.success(request, f"Molde '{nombre_molde}' actualizado correctamente.")
    return redirect('paros:catalogo_moldes')
